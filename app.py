"""SCORPION ENGINE V4 PRO — Sistema completo"""
import streamlit as st
import os, json, re, time, math, io, base64, sqlite3, hashlib, warnings, requests, random
from bs4 import BeautifulSoup
from datetime import date, timedelta, time as dtime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════
API_FOOTBALL_KEY = os.getenv("API_FOOTBALL_KEY", "124c9519df145caf883cd82f0b2a4671")
ADMIN_PWD        = os.getenv("ADMIN_PASSWORD",   "scorpion_admin_2025")
DB_PATH          = "/tmp/scorpion_v4.db"

LIGAS = {
    "🏴󠁧󠁢󠁥󠁮󠁧󠁿 Premier League":  39,
    "🇪🇸 La Liga":           140,
    "🇩🇪 Bundesliga":        78,
    "🇮🇹 Serie A":           135,
    "🇫🇷 Ligue 1":           61,
    "🇳🇱 Eredivisie":        88,
    "🇵🇹 Primeira Liga":     94,
    "🌍 Champions League":   2,
    "🌍 Europa League":      3,
    "🌍 Conference League":  848,
    "🌎 Libertadores":       13,
    "🌎 Sudamericana":       11,
    "🌎 Copa America":       9,
    "🌍 Mundial FIFA 2026":  1,
    "🇺🇸 MLS":               253,
    "🇲🇽 Liga MX":           262,
    "🇨🇴 Liga BetPlay":      239,
    "🇦🇷 Primera Division":  128,
    "🇧🇷 Brasileirao":       71,
    "🇹🇷 Super Lig":         203,
    "🇸🇦 Saudi Pro League":  307,
}

PROM_LIGA = {
    "premier":{"gm":1.54,"gc":1.11,"tiros":14.2,"corners":5.2,"tarj":1.8},
    "la liga":{"gm":1.62,"gc":1.08,"tiros":13.8,"corners":5.5,"tarj":2.2},
    "bundesliga":{"gm":1.82,"gc":1.28,"tiros":15.1,"corners":5.8,"tarj":1.6},
    "serie a":{"gm":1.48,"gc":1.07,"tiros":13.2,"corners":5.0,"tarj":2.4},
    "ligue":{"gm":1.51,"gc":1.07,"tiros":13.5,"corners":5.1,"tarj":2.0},
    "libertadores":{"gm":1.32,"gc":1.08,"tiros":12.0,"corners":4.8,"tarj":2.8},
    "sudamericana":{"gm":1.28,"gc":1.07,"tiros":11.8,"corners":4.7,"tarj":2.9},
    "eredivisie":{"gm":1.88,"gc":1.32,"tiros":15.5,"corners":5.9,"tarj":1.5},
    "mls":{"gm":1.45,"gc":1.20,"tiros":13.0,"corners":5.0,"tarj":1.9},
    "colombia":{"gm":1.25,"gc":1.10,"tiros":11.5,"corners":4.6,"tarj":2.6},
    "mundial":{"gm":1.35,"gc":1.05,"tiros":13.0,"corners":4.8,"tarj":1.8},
    "world cup":{"gm":1.35,"gc":1.05,"tiros":13.0,"corners":4.8,"tarj":1.8},
    "copa america":{"gm":1.28,"gc":1.08,"tiros":12.2,"corners":4.7,"tarj":2.2},
    "champions":{"gm":1.45,"gc":1.05,"tiros":14.0,"corners":5.2,"tarj":1.7},
    "europa":{"gm":1.42,"gc":1.08,"tiros":13.5,"corners":5.0,"tarj":1.9},
    "default":{"gm":1.40,"gc":1.15,"tiros":12.8,"corners":4.9,"tarj":2.0},
}

LIGAS_TOP = ["premier","la liga","bundesliga","serie a","ligue","champions",
             "europa","libertadores","mundial","world cup","copa america","eredivisie"]

UNDERSTAT_MAP = {
    "premier":"EPL","premier league":"EPL","la liga":"La_liga",
    "bundesliga":"Bundesliga","serie a":"Serie_A","ligue 1":"Ligue_1","ligue":"Ligue_1",
}

TORNEOS_FIFA = ["mundial","world cup","fifa","copa del mundo","nations league",
                "eurocopa","euro 20","copa america","gold cup","afcon"]

SELECCIONES = {
    "argentina","brasil","brazil","francia","france","alemania","germany",
    "espana","spain","portugal","inglaterra","england","italia","italy",
    "belgica","belgium","croacia","croatia","holanda","netherlands","uruguay",
    "colombia","chile","mexico","estados unidos","usa","japon","japan",
    "marruecos","morocco","senegal","australia","corea","korea","suiza",
    "switzerland","dinamarca","denmark","polonia","poland","austria",
    "turquia","turkey","serbia","ecuador","peru","ghana","nigeria","camerun",
    "sudafrica","south africa","arabia","saudi","iran","qatar","canada",
    "gales","wales","escocia","scotland","hungria","georgia","eslovenia",
}

SH = requests.Session()
SH.headers.update({"User-Agent":"Mozilla/5.0","Accept":"application/json,text/html,*/*"})

# ══════════════════════════════════════════════════════════
# BASE DE DATOS
# ══════════════════════════════════════════════════════════
def get_conn():
    c = sqlite3.connect(DB_PATH, check_same_thread=False)
    c.row_factory = sqlite3.Row
    return c

def init_db():
    c = get_conn()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS usuarios (
        cedula TEXT PRIMARY KEY, nombre TEXT, plan TEXT DEFAULT 'gratis',
        fecha_inicio TEXT, dias INTEGER DEFAULT 36500, activo INTEGER DEFAULT 1,
        es_admin INTEGER DEFAULT 0, pwd_hash TEXT,
        consultas_hoy INTEGER DEFAULT 0, fecha_reset TEXT,
        creado TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS picks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT, liga TEXT, local TEXT, visitante TEXT, hora TEXT,
        mercado TEXT, detalle TEXT, cuota REAL, edge REAL,
        confianza REAL, rango TEXT, resultado TEXT, acierto INTEGER,
        notas TEXT, plan_min TEXT DEFAULT 'gratis', auto INTEGER DEFAULT 0,
        creado TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS historial (
        id INTEGER PRIMARY KEY AUTOINCREMENT, cedula TEXT, fecha TEXT,
        liga TEXT, local TEXT, visitante TEXT,
        p1 REAL, px REAL, p2 REAL, xg_l REAL, xg_v REAL,
        over25 REAL, btts REAL, mercado TEXT, rango TEXT, confianza REAL,
        creado TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS escalera (
        id INTEGER PRIMARY KEY AUTOINCREMENT, fecha TEXT,
        pasos TEXT, estado TEXT DEFAULT 'activa',
        creado TEXT DEFAULT CURRENT_TIMESTAMP
    );
    """)
    h = hashlib.sha256(ADMIN_PWD.encode()).hexdigest()
    c.execute("INSERT OR IGNORE INTO usuarios (cedula,nombre,plan,fecha_inicio,dias,activo,es_admin,pwd_hash) VALUES (?,?,?,?,?,?,?,?)",
              ("admin","Administrador","admin",str(date.today()),36500,1,1,h))
    c.commit(); c.close()

def db_get(cedula):
    c=get_conn(); r=c.execute("SELECT * FROM usuarios WHERE cedula=?",(cedula,)).fetchone(); c.close()
    return dict(r) if r else None

def db_todos():
    c=get_conn(); r=c.execute("SELECT * FROM usuarios ORDER BY creado DESC").fetchall(); c.close()
    return [dict(x) for x in r]

def db_guardar_usuario(cedula,nombre,plan,dias,fi,activo=1):
    c=get_conn()
    c.execute("""INSERT OR REPLACE INTO usuarios
        (cedula,nombre,plan,fecha_inicio,dias,activo,es_admin,pwd_hash,creado)
        VALUES(?,?,?,?,?,?,
        COALESCE((SELECT es_admin FROM usuarios WHERE cedula=?),0),
        COALESCE((SELECT pwd_hash FROM usuarios WHERE cedula=?),NULL),
        COALESCE((SELECT creado FROM usuarios WHERE cedula=?),CURRENT_TIMESTAMP))""",
        (cedula,nombre,plan,str(fi),int(dias),int(activo),cedula,cedula,cedula))
    c.commit(); c.close()

def db_acceso(cedula):
    u=db_get(cedula)
    if not u: return False,"no_existe",0
    if not u["activo"]: return False,"inactivo",0
    if u["es_admin"]: return True,"admin",99999
    if u["plan"]=="gratis": return True,"gratis",99999
    inicio=date.fromisoformat(u["fecha_inicio"])
    vence=inicio+timedelta(days=u["dias"]); dr=(vence-date.today()).days
    if date.today()>vence: return False,"vencido",0
    return True,u["plan"],dr

def db_login_admin(pwd):
    u=db_get("admin")
    return u and u.get("pwd_hash")==hashlib.sha256(pwd.encode()).hexdigest()

def db_consultas(cedula):
    c=get_conn(); hoy=str(date.today())
    u=c.execute("SELECT consultas_hoy,fecha_reset FROM usuarios WHERE cedula=?",(cedula,)).fetchone()
    if u:
        if u["fecha_reset"]!=hoy:
            c.execute("UPDATE usuarios SET consultas_hoy=1,fecha_reset=? WHERE cedula=?",(hoy,cedula))
            c.commit(); c.close(); return 1
        n=(u["consultas_hoy"] or 0)+1
        c.execute("UPDATE usuarios SET consultas_hoy=? WHERE cedula=?",(n,cedula))
        c.commit(); c.close(); return n
    c.close(); return 0

def db_pick_guardar(fecha,liga,local,visitante,hora,mercado,detalle,cuota,edge,confianza,rango,notas,plan_min,auto=0):
    c=get_conn()
    c.execute("INSERT INTO picks (fecha,liga,local,visitante,hora,mercado,detalle,cuota,edge,confianza,rango,notas,plan_min,auto) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
              (fecha,liga,local,visitante,hora,mercado,detalle,cuota,edge,confianza,rango,notas,plan_min,auto))
    c.commit(); c.close()

def db_picks_get(fecha=None, plan="gratis"):
    c=get_conn()
    orden={"gratis":0,"dia":1,"semana":2,"mes":3,"admin":4}
    nivel=orden.get(plan,0)
    if fecha:
        rows=c.execute("SELECT * FROM picks WHERE fecha=? ORDER BY confianza DESC",(fecha,)).fetchall()
    else:
        rows=c.execute("SELECT * FROM picks ORDER BY fecha DESC,confianza DESC LIMIT 100").fetchall()
    c.close()
    resultado=[]
    for r in [dict(x) for x in rows]:
        nivel_min=orden.get(r.get("plan_min","gratis"),0)
        if nivel<nivel_min:
            r["mercado"]=f"🔒 Plan {r['plan_min'].upper()}"
            r["detalle"]=""; r["cuota"]=None; r["edge"]=None; r["confianza"]=None
        resultado.append(r)
    return resultado

def db_historial_guardar(cedula,p,calc):
    c=get_conn()
    c.execute("INSERT INTO historial (cedula,fecha,liga,local,visitante,p1,px,p2,xg_l,xg_v,over25,btts,mercado,rango,confianza) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
              (cedula,p.get("dia",str(date.today())),p.get("liga",""),p.get("local",""),p.get("visitante",""),
               calc.get("p1"),calc.get("px"),calc.get("p2"),calc.get("xl"),calc.get("xv"),
               calc.get("over25"),calc.get("btts_si"),calc.get("mk2",""),calc.get("rango",""),calc.get("confianza")))
    c.commit(); c.close()

def db_historial_get(cedula,limite=30):
    c=get_conn()
    r=c.execute("SELECT * FROM historial WHERE cedula=? ORDER BY creado DESC LIMIT ?",(cedula,limite)).fetchall()
    c.close(); return [dict(x) for x in r]

def db_stats():
    c=get_conn()
    tu=c.execute("SELECT COUNT(*) FROM usuarios WHERE es_admin=0").fetchone()[0]
    tp=c.execute("SELECT COUNT(*) FROM picks").fetchone()[0]
    th=c.execute("SELECT COUNT(*) FROM historial").fetchone()[0]
    c.close()
    activos=sum(1 for u in db_todos() if u.get("es_admin")==0 and db_acceso(u["cedula"])[0])
    return {"usuarios":tu,"activos":activos,"picks":tp,"historial":th}

# ══════════════════════════════════════════════════════════
# MOTOR MATEMATICO
# ══════════════════════════════════════════════════════════
def pp(lam,k):
    if lam<=0 or k<0: return 0.0
    return (math.exp(-lam)*(lam**k))/math.factorial(min(k,20))

def dc_tau(x,y,xl,xv,rho=-0.1):
    if x==0 and y==0: return 1-xl*xv*rho
    if x==1 and y==0: return 1+xv*rho
    if x==0 and y==1: return 1+xl*rho
    if x==1 and y==1: return 1-rho
    return 1.0

def poisson_1x2(xl,xv):
    p1=px=p2=0.0
    for i in range(9):
        for j in range(9):
            p=pp(xl,i)*pp(xv,j)
            if i>j: p1+=p
            elif i==j: px+=p
            else: p2+=p
    return round(p1*100,1),round(px*100,1),round(p2*100,1)

def dc_1x2(xl,xv,rho=-0.1):
    m={}
    for i in range(9):
        for j in range(9):
            m[(i,j)]=max(0,pp(xl,i)*pp(xv,j)*dc_tau(i,j,xl,xv,rho))
    t=sum(m.values())
    if t>0: m={k:v/t for k,v in m.items()}
    p1=sum(v for (i,j),v in m.items() if i>j)
    px=sum(v for (i,j),v in m.items() if i==j)
    p2=sum(v for (i,j),v in m.items() if i<j)
    return round(p1*100,1),round(px*100,1),round(p2*100,1)

def monte_carlo(xl,xv,n=3000):
    v1=vx=v2=0; gt=[]; mk={}
    for _ in range(n):
        gl=gv=0; a=0.0; u=random.random()
        for k in range(15):
            a+=pp(xl,k)
            if u<=a: gl=k; break
        a=0.0; u=random.random()
        for k in range(15):
            a+=pp(xv,k)
            if u<=a: gv=k; break
        if gl>gv: v1+=1
        elif gl==gv: vx+=1
        else: v2+=1
        gt.append(gl+gv)
        mk[f"{gl}-{gv}"]=mk.get(f"{gl}-{gv}",0)+1
    top=dict(sorted(mk.items(),key=lambda x:x[1],reverse=True)[:6])
    top={k:round(v/n*100,1) for k,v in top.items()}
    return {"p1":round(v1/n*100,1),"px":round(vx/n*100,1),"p2":round(v2/n*100,1),
            "avg":round(sum(gt)/n,2),"o25":round(sum(1 for g in gt if g>2)/n*100),"top":top}

def elo_1x2(elo_l,elo_v,vent=50):
    e=1/(1+10**((elo_v-(elo_l+vent))/400))
    p1=round(e*100,1); p2=round((1-e)*100,1); px=max(0,round(100-p1-p2,1))
    return p1,px,p2

def get_prom(liga):
    l=liga.lower()
    for k in PROM_LIGA:
        if k in l: return PROM_LIGA[k]
    return PROM_LIGA["default"]

def es_top(liga):
    l=liga.lower()
    return any(k in l for k in LIGAS_TOP)

def calcular(gml,gcl,gmv,gcv,liga,elo_l=None,elo_v=None):
    pr=get_prom(liga)
    if gml and gcv:  xl=round(gml*(gcv/pr["gc"])*1.08,2)
    elif gml:        xl=round(gml*1.08,2)
    else:            xl=round(pr["gm"]*1.08,2)
    if gmv and gcl:  xv=round(gmv*(gcl/pr["gc"]),2)
    elif gmv:        xv=round(gmv,2)
    else:            xv=round(pr["gm"]*0.78,2)
    if elo_l and elo_v:
        f=1+(elo_l-elo_v)/4000
        xl=round(xl*min(max(f,0.7),1.4),2)
        xv=round(xv*min(max(1/f,0.7),1.4),2)
    xl=max(0.15,xl); xv=max(0.10,xv); xt=round(xl+xv,2)
    p1_po,px_po,p2_po=poisson_1x2(xl,xv)
    p1_dc,px_dc,p2_dc=dc_1x2(xl,xv)
    mc=monte_carlo(xl,xv)
    p1_mc,px_mc,p2_mc=mc["p1"],mc["px"],mc["p2"]
    if elo_l and elo_v: p1_el,px_el,p2_el=elo_1x2(elo_l,elo_v)
    else: p1_el,px_el,p2_el=p1_po,px_po,p2_po
    p1=round(p1_po*.35+p1_dc*.30+p1_mc*.20+p1_el*.15,1)
    px=round(px_po*.35+px_dc*.30+px_mc*.20+px_el*.15,1)
    p2=round(max(0,100-p1-px),1)
    conv=100-(abs(p1_po-p1_dc)*.4+abs(p1_po-p1_mc)*.3+abs(p1_po-p1_el)*.3)
    conf=round(max(0,min(100,conv)))
    datos_r=gml is not None and gmv is not None
    if datos_r and conf>=75 and es_top(liga): rango="A+"
    elif datos_r and conf>=55: rango="B"
    elif conf>=40: rango="C"
    else: rango="D"
    p0=pp(xt,0);p1_=pp(xt,1);p2_=pp(xt,2);p3_=pp(xt,3)
    o05=round((1-p0)*100); o15=round((1-p0-p1_)*100)
    o25=round((1-p0-p1_-p2_)*100); o35=round((1-p0-p1_-p2_-p3_)*100)
    btts_si=round((1-pp(xl,0))*(1-pp(xv,0))*100); btts_no=100-btts_si
    # Corners: promedio real es ~10 por partido, formula calibrada
    cmu=round(min(14, xt*1.4+6.5),1); sc=2.0
    def poc(l):
        z=(l-cmu)/sc; return max(5,min(95,round((1-0.5*(1+math.erf(z/math.sqrt(2))))*100)))
    c75=poc(7.5);c85=poc(8.5);c95=poc(9.5);c105=poc(10.5)
    dif=abs(p1-p2)
    base=3.8 if any(k in liga.lower() for k in ["la liga","serie a","ligue","libertadores","colombia"]) else 3.2
    tmu=round(base+(1-dif/60)*1.8,1)
    # Tiros al arco estimados — proporcional al xG real, no multiplicado
    # Promedio real: ~1 tiro al arco cada 0.18 xG aproximadamente
    tiros_l   = round(max(2.0, xl * 3.8), 1)   # xG * factor calibrado
    tiros_v   = round(max(1.5, xv * 3.8), 1)
    tiros_tot = round(tiros_l + tiros_v, 1)
    mk="1" if p1>p2 and p1>px else ("X" if px>=p1 and px>=p2 else "2")
    if p1>p2 and p1>px and o25>=55: mk2="Gana Local + Over 1.5"
    elif p2>p1 and p2>px and o25>=55: mk2="Gana Visita + Over 1.5"
    elif o25>=65: mk2="Over 2.5 Goles"
    elif btts_si>=62: mk2="BTTS — Ambos Marcan"
    elif c95>=65: mk2="Corners Over 9.5"
    elif p1>p2 and p1>px: mk2="Victoria Local"
    elif p2>p1 and p2>px: mk2="Victoria Visitante"
    else: mk2="Empate posible"
    top_ex={}
    for i in range(7):
        for j in range(7):
            pij=round(pp(xl,i)*pp(xv,j)*100,1)
            if pij>=0.5: top_ex[f"{i}-{j}"]=pij
    top_ex=dict(sorted(top_ex.items(),key=lambda x:x[1],reverse=True)[:8])
    # Mercados completos para picks
    # Cuotas de referencia calibradas con margen de casa tipico (5-8%)
    # Cuota justa = 100/prob, cuota mercado = cuota_justa * 1.07 (margen)
    def cuota_ref(prob, cuota_minima=1.05):
        if prob <= 0: return None
        cj = 100 / prob  # cuota justa sin margen
        cm = round(cj * 1.07, 2)  # con margen de casa 7%
        return max(cuota_minima, cm)

    mercados_picks = [
        ("Over 0.5 Goles",       o05,     cuota_ref(o05)),
        ("Over 1.5 Goles",       o15,     cuota_ref(o15)),
        ("Over 2.5 Goles",       o25,     cuota_ref(o25)),
        ("Over 3.5 Goles",       o35,     cuota_ref(o35)),
        ("BTTS — Ambos Marcan",  btts_si, cuota_ref(btts_si)),
        ("BTTS — No",            btts_no, cuota_ref(btts_no)),
        ("Corners +7.5",         c75,     cuota_ref(c75)),
        ("Corners +8.5",         c85,     cuota_ref(c85)),
        ("Corners +9.5",         c95,     cuota_ref(c95)),
        ("Corners +10.5",        c105,    cuota_ref(c105)),
        ("Tarjetas +1.5",        round(min(85,(tmu/4.5)*100),0), cuota_ref(round(min(85,(tmu/4.5)*100),0))),
        ("Tarjetas +2.5",        round(min(70,(tmu/5.5)*100),0), cuota_ref(round(min(70,(tmu/5.5)*100),0))),
        ("Tarjetas +3.5",        round(min(50,(tmu/7)*100),0),   cuota_ref(round(min(50,(tmu/7)*100),0))),
    ]
    # Resultado 1X2 — cuota justa desde probabilidades del modelo
    if p1>px and p1>p2:
        mercados_picks.insert(0,("Victoria Local (1)", p1, cuota_ref(p1)))
    elif p2>px and p2>p1:
        mercados_picks.insert(0,("Victoria Visitante (2)", p2, cuota_ref(p2)))
    else:
        mercados_picks.insert(0,("Empate (X)", px, cuota_ref(px)))
    return {
        "xl":xl,"xv":xv,"xt":xt,
        "p1_po":p1_po,"px_po":px_po,"p2_po":p2_po,
        "p1_dc":p1_dc,"px_dc":px_dc,"p2_dc":p2_dc,
        "p1_mc":p1_mc,"px_mc":px_mc,"p2_mc":p2_mc,
        "p1_el":p1_el,"px_el":px_el,"p2_el":p2_el,
        "p1":p1,"px":px,"p2":p2,"confianza":conf,"rango":rango,
        "o05":o05,"over15":o15,"over25":o25,"over35":o35,
        "btts_si":btts_si,"btts_no":btts_no,
        "cmu":cmu,"c75":c75,"c85":c85,"c95":c95,"c105":c105,
        "corners_str":f"~{int(cmu)} (+9.5:{c95}% | +8.5:{c85}%)",
        "tiros_l":tiros_l,"tiros_v":tiros_v,"tiros_tot":tiros_tot,
        "tmu":tmu,"tar_str":f"~{max(2,int(tmu)-1)}-{int(tmu)+1} tarjetas",
        "mk":mk,"mk2":mk2,"top_ex":top_ex,
        "avg_g":mc["avg"],"top_mc":mc["top"],"datos_r":datos_r,
        "mercados_picks":mercados_picks,
    }

# ══════════════════════════════════════════════════════════
# API-FOOTBALL
# ══════════════════════════════════════════════════════════
_tc={}
def get_temp(lid):
    if lid in _tc: return _tc[lid]
    fallback={1:2026,253:2025,262:2025,239:2025,71:2025}
    try:
        r=SH.get("https://v3.football.api-sports.io/leagues",
                  headers={"x-apisports-key":API_FOOTBALL_KEY},
                  params={"id":lid,"current":"true"},timeout=10)
        data=r.json().get("response",[])
        if data:
            seasons=data[0].get("seasons",[])
            activa=[s for s in seasons if s.get("current")]
            if activa: t=activa[0]["year"]; _tc[lid]=t; return t
            if seasons: t=seasons[-1]["year"]; _tc[lid]=t; return t
    except: pass
    t=fallback.get(lid,2024); _tc[lid]=t; return t

def get_fx_dia(lid,fecha):
    t=get_temp(lid); h={"x-apisports-key":API_FOOTBALL_KEY}
    for p in [{"league":lid,"season":t,"date":fecha},{"league":lid,"date":fecha}]:
        try:
            r=SH.get("https://v3.football.api-sports.io/fixtures",headers=h,params=p,timeout=15)
            d=r.json().get("response",[])
            if d: return d
        except: pass
    return []

def get_fx_rango(lid,desde,hasta):
    t=get_temp(lid); h={"x-apisports-key":API_FOOTBALL_KEY}
    for p in [{"league":lid,"season":t,"from":desde,"to":hasta},{"league":lid,"from":desde,"to":hasta}]:
        try:
            r=SH.get("https://v3.football.api-sports.io/fixtures",headers=h,params=p,timeout=15)
            d=r.json().get("response",[])
            if d: return d
        except: pass
    return []

def fx2p(f):
    dt=f["fixture"]["date"]
    return {"dia":dt[:10],"hora":dt[11:16],"hora_sort":int(dt[11:13])*60+int(dt[14:16]),
            "liga":f["league"]["name"],"liga_id":f["league"]["id"],
            "local":f["teams"]["home"]["name"],"visitante":f["teams"]["away"]["name"],
            "tid_l":f["teams"]["home"]["id"],"tid_v":f["teams"]["away"]["id"]}

_sc={}
def get_stats_api(tid,lid):
    k=f"{tid}_{lid}"
    if k in _sc: return _sc[k]
    t=get_temp(lid)
    try:
        r=SH.get("https://v3.football.api-sports.io/teams/statistics",
                  headers={"x-apisports-key":API_FOOTBALL_KEY},
                  params={"team":tid,"season":t,"league":lid},timeout=12)
        d=r.json().get("response",{})
        gf=d.get("goals",{}).get("for",{}).get("average",{}).get("total")
        ga=d.get("goals",{}).get("against",{}).get("average",{}).get("total")
        res=(float(gf) if gf else None,float(ga) if ga else None)
        _sc[k]=res; return res
    except: pass
    _sc[k]=(None,None); return None,None

_uc={}
def get_understat(eq,liga,temp=2024):
    lu=next((v for k,v in UNDERSTAT_MAP.items() if k in liga.lower()),None)
    if not lu: return None,None,None,None
    ck=f"{lu}_{temp}"
    if ck not in _uc:
        try:
            r=SH.get(f"https://understat.com/league/{lu}/{temp}",timeout=15)
            for sc in BeautifulSoup(r.text,"lxml").find_all("script"):
                if "teamsData" in str(sc):
                    m=re.search(r"JSON\.parse\(.(.*?)\.replace",str(sc))
                    if m: _uc[ck]=json.loads(m.group(1).encode().decode("unicode_escape")); break
        except: _uc[ck]={}
    data=_uc.get(ck,{})
    el=eq.lower()
    for tn,stats in data.items():
        if any(p in tn.lower() for p in el.split()[:2]):
            h=stats.get("history",[])[-10:]
            if h:
                return (round(sum(x.get("xG",0) for x in h)/len(h),2),
                        round(sum(x.get("xGA",0) for x in h)/len(h),2),
                        round(sum(x.get("scored",0) for x in h)/len(h),2),
                        round(sum(x.get("missed",0) for x in h)/len(h),2))
    return None,None,None,None

def limpiar_nombre(nombre):
    """Limpia sufijos de ciudad/pais para mejorar busqueda: Botafogo-RJ -> Botafogo"""
    import re as re2
    # Quitar sufijos comunes: -RJ, -SP, -MG, -PR, etc.
    nombre = re2.sub(r"-\s*[A-Z]{2,3}$", "", nombre.strip())
    # Quitar parentesis: Atletico (MG) -> Atletico
    nombre = re2.sub(r"\s*\([^)]+\)", "", nombre)
    # Quitar puntos y guiones extra
    nombre = nombre.strip(" .-")
    return nombre

def get_tsdb(nombre):
    nombre_clean = limpiar_nombre(nombre)
    for buscar in [nombre_clean, nombre]:
        try:
            r=SH.get(f"https://www.thesportsdb.com/api/v1/json/3/searchteams.php?t={requests.utils.quote(buscar)}",timeout=8)
            d=r.json()
            if d and d.get("teams"): return d["teams"][0]
        except: pass
    return None

def buscar_equipo_api_football(nombre, liga_id=None):
    """Busca el team_id en API-Football por nombre."""
    nombre_clean = limpiar_nombre(nombre)
    h = {"x-apisports-key": API_FOOTBALL_KEY}
    for buscar in [nombre_clean, nombre]:
        try:
            params = {"name": buscar}
            if liga_id: params["league"] = liga_id
            r = SH.get("https://v3.football.api-sports.io/teams",
                       headers=h, params=params, timeout=10)
            data = r.json().get("response", [])
            if data:
                return data[0]["team"]["id"], data[0]["team"]["name"]
        except: pass
    return None, None

def stats_tsdb(tid):
    try:
        r=SH.get(f"https://www.thesportsdb.com/api/v1/json/3/eventslast.php?id={tid}",timeout=10)
        d=r.json()
        if d and d.get("results"):
            # Filtrar solo partidos con score valido
            raw=[x for x in d["results"] if x.get("intHomeScore") is not None]
            raw=raw[-10:]; gml=[]; gcl=[]
            for x in raw:
                try:
                    sh=int(x.get("intHomeScore") or 0); sa=int(x.get("intAwayScore") or 0)
                    if str(x.get("idHomeTeam",""))==str(tid): gml.append(sh);gcl.append(sa)
                    else: gml.append(sa);gcl.append(sh)
                except: pass
            if gml: return round(sum(gml)/len(gml),2),round(sum(gcl)/len(gcl),2)
    except: pass
    return None,None

_ec={}
def get_elo(eq):
    if eq in _ec: return _ec[eq]
    try:
        r=SH.get(f"http://api.clubelo.com/{eq.replace(' ','-').replace('.','')}",timeout=8)
        if r.status_code==200 and len(r.text)>50:
            ul=r.text.strip().split("\n")[-1].split(",")
            if len(ul)>=4: elo=float(ul[3]); _ec[eq]=elo; return elo
    except: pass
    _ec[eq]=None; return None

def obtener_stats_detalle(nombre, liga, tid=None, lid=None):
    """
    Obtiene stats detallados desde multiples fuentes en cascada:
    1. API-Football (goles, stats de temporada) — requiere API key
    2. Understat (xG reales, ligas europeas top)
    3. TheSportsDB (ultimos partidos, cualquier liga/seleccion)
    4. ClubElo (ELO rating europeo)
    """
    det = {
        "nombre": nombre, "liga": liga, "fuente": "Sin datos", "ok": False,
        "gm": None, "gc": None, "elo": None,
        "xg": None, "xga": None,
        "tiros_pg": None, "tarj_pg": None, "posesion": None,
        "ultimos5": [], "ganados5": 0, "empatados5": 0, "perdidos5": 0,
        "goles_fav5": 0, "goles_con5": 0,
        "fuentes_usadas": [],
    }
    torneo = any(k in liga.lower() for k in TORNEOS_FIFA)
    selec  = any(k in nombre.lower() for k in SELECCIONES)

    # ── FUENTE 1: API-Football (stats completos de temporada) ────────────
    # Si no tenemos team_id, buscarlo por nombre
    if not tid and lid:
        tid_found, _ = buscar_equipo_api_football(nombre, lid)
        if tid_found: tid = tid_found
    if not tid:
        tid_found, _ = buscar_equipo_api_football(nombre)
        if tid_found: tid = tid_found

    if tid and lid:
        try:
            t = get_temp(lid)
            h = {"x-apisports-key": API_FOOTBALL_KEY}
            r = SH.get("https://v3.football.api-sports.io/teams/statistics",
                       headers=h, params={"team":tid,"season":t,"league":lid}, timeout=12)
            d = r.json().get("response", {})
            if d:
                gf = d.get("goals",{}).get("for",{}).get("average",{}).get("total")
                ga = d.get("goals",{}).get("against",{}).get("average",{}).get("total")
                shots = d.get("shots",{}).get("total",{}).get("average")
                cards = d.get("cards",{}).get("yellow",{})
                # Sumar tarjetas amarillas por intervalo
                tarj = 0
                if isinstance(cards, dict):
                    for v in cards.values():
                        if isinstance(v, dict) and v.get("total"):
                            try: tarj += int(v["total"])
                            except: pass
                played = d.get("fixtures",{}).get("played",{}).get("total") or 1
                if gf:
                    det.update({
                        "gm": round(float(gf),2),
                        "gc": round(float(ga),2) if ga else None,
                        "tiros_pg": round(float(shots),1) if shots else None,
                        "tarj_pg": round(tarj/played,2) if tarj and played else None,
                        "fuente": "API-Football",
                        "ok": True,
                    })
                    det["fuentes_usadas"].append("API-Football")
        except: pass

    # ── FUENTE 2: Understat (xG reales, ligas europeas) ─────────────────
    if not torneo and not selec:
        try:
            xg, xga, gm_u, gc_u = get_understat(nombre, liga)
            if xg is not None:
                det["xg"]  = xg
                det["xga"] = xga
                if not det["ok"]:
                    det.update({"gm": gm_u, "gc": gc_u, "fuente": "Understat", "ok": True})
                det["fuentes_usadas"].append("Understat")
            time.sleep(0.2)
        except: pass

    # ── FUENTE 3: TheSportsDB (ultimos partidos + stats basicos) ─────────
    try:
        td = get_tsdb(nombre)
        if td:
            team_id = td.get("idTeam","")
            if not det["ok"]:
                gm2, gc2 = stats_tsdb(team_id)
                if gm2:
                    det.update({"gm": gm2, "gc": gc2, "fuente": "TheSportsDB", "ok": True})
            det["fuentes_usadas"].append("TheSportsDB")
            # Ultimos partidos con detalle de resultado
            try:
                r2 = SH.get(f"https://www.thesportsdb.com/api/v1/json/3/eventslast.php?id={team_id}", timeout=10)
                d2 = r2.json()
                if d2 and d2.get("results"):
                    ultimos = [x for x in d2["results"] if x.get("intHomeScore") is not None][-5:]
                    for x in ultimos:
                        try:
                            sh = int(x.get("intHomeScore") or 0)
                            sa = int(x.get("intAwayScore") or 0)
                            es_l = str(x.get("idHomeTeam","")) == str(team_id)
                            gf_p = sh if es_l else sa
                            gc_p = sa if es_l else sh
                            rival = x.get("strAwayTeam" if es_l else "strHomeTeam","?")
                            liga_p = x.get("strLeague","")
                            if gf_p > gc_p:    res="G"; det["ganados5"]+=1
                            elif gf_p == gc_p: res="E"; det["empatados5"]+=1
                            else:              res="P"; det["perdidos5"]+=1
                            det["goles_fav5"] += gf_p
                            det["goles_con5"] += gc_p
                            det["ultimos5"].append({
                                "rival": rival[:16], "res": res,
                                "goles": f"{gf_p}-{gc_p}",
                                "local": "Local" if es_l else "Visita",
                                "liga": liga_p[:18],
                            })
                        except: pass
            except: pass
        time.sleep(0.2)
    except: pass

    # ── FUENTE 4: ClubElo (ELO rating) ───────────────────────────────────
    try:
        elo = get_elo(nombre)
        if elo:
            det["elo"] = elo
            det["fuentes_usadas"].append("ClubElo")
    except: pass

    return det

def obtener_stats(nombre,liga,tid=None,lid=None):
    nombre_c = limpiar_nombre(nombre)
    d = obtener_stats_detalle(nombre_c, liga, tid, lid)
    if not d["ok"] and nombre_c != nombre:
        d = obtener_stats_detalle(nombre, liga, tid, lid)
    return {"gm":d["gm"],"gc":d["gc"],"elo":d["elo"],"fuente":d["fuente"],"ok":d["ok"]}

def leer_imagen(img_bytes,mt="image/jpeg"):
    try:
        b64=base64.standard_b64encode(img_bytes).decode()
        payload={"model":"claude-sonnet-4-6","max_tokens":1500,"messages":[{"role":"user","content":[
            {"type":"image","source":{"type":"base64","media_type":mt,"data":b64}},
            {"type":"text","text":"Analiza esta imagen de fixtures de futbol. Devuelve SOLO JSON valido sin markdown: {\"partidos\":[{\"hora\":\"19:00\",\"liga\":\"Premier League\",\"local\":\"Arsenal\",\"visitante\":\"Chelsea\"}]}. Si no hay hora usa '00:00'. Solo partidos claramente legibles."}
        ]}]}
        r=requests.post("https://api.anthropic.com/v1/messages",
                        headers={"Content-Type":"application/json"},json=payload,timeout=30)
        text=r.json().get("content",[{}])[0].get("text","")
        text=re.sub(r"```json|```","",text).strip()
        data=json.loads(text); result=[]
        for p in data.get("partidos",[]):
            hora=str(p.get("hora","00:00"))
            try: h,m=hora.split(":"); hs=int(h)*60+int(m)
            except: hs=0
            if p.get("local") and p.get("visitante"):
                result.append({"dia":str(date.today()),"hora":hora,"hora_sort":hs,
                               "liga":p.get("liga","Sin Liga"),"liga_id":0,
                               "local":p["local"],"visitante":p["visitante"],"tid_l":None,"tid_v":None})
        return result
    except Exception as e:
        st.error(f"Error leyendo imagen: {e}"); return []

def leer_excel(fb):
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(fb),data_only=True); ws=wb.active
    partidos=[]; cur_liga=""
    TORN=["copa","liga","premier","bundesliga","serie","ligue","championship","apertura","clausura",
          "europa","champions","sudamericana","libertadores","mls","eredivisie","primera","superliga",
          "mundial","world cup","fifa","copa del mundo","nations","eurocopa","copa america"]
    def fix(s):
        s=str(s).strip().replace("\xa0","")
        for n in range(3,len(s)//2+1):
            if s[:n]==s[n:n*2]: return s[:n]
        return s
    for ri in range(1,ws.max_row+1):
        val=ws.cell(row=ri,column=1).value
        if val is None: continue
        if isinstance(val,dtime):
            try:
                l=fix(ws.cell(row=ri+1,column=1).value or "")
                v=fix(ws.cell(row=ri+2,column=1).value or "")
                if l not in ("","--") and v not in ("","--"):
                    partidos.append({"dia":str(date.today()),"hora":f"{val.hour:02d}:{val.minute:02d}",
                                     "hora_sort":val.hour*60+val.minute,"liga":cur_liga or "Sin Liga",
                                     "liga_id":0,"local":l,"visitante":v,"tid_l":None,"tid_v":None})
            except: pass
            continue
        vs=str(val).strip().replace("\xa0","")
        if any(k in vs.lower() for k in TORN) and len(vs)<80: cur_liga=vs
    return partidos

def analizar_lista(partidos,usar_api=True,prog=None):
    resultados=[]; n=len(partidos)
    for idx,p in enumerate(partidos):
        gml=gcl=gmv=gcv=elo_l=elo_v=None; fl=fv="Prom.liga"
        if usar_api:
            sl=obtener_stats(p["local"],p["liga"],p.get("tid_l"),p.get("liga_id"))
            sv=obtener_stats(p["visitante"],p["liga"],p.get("tid_v"),p.get("liga_id"))
            if sl["ok"]: gml=sl["gm"];gcl=sl["gc"];elo_l=sl["elo"];fl=sl["fuente"]
            if sv["ok"]: gmv=sv["gm"];gcv=sv["gc"];elo_v=sv["elo"];fv=sv["fuente"]
        calc=calcular(gml,gcl,gmv,gcv,p["liga"],elo_l,elo_v)
        resultados.append({**p,**calc,"fuente_l":fl,"fuente_v":fv})
        if prog: prog.progress((idx+1)/n,text=f"{idx+1}/{n}: {p['local']} vs {p['visitante']}")
    return resultados

# ══════════════════════════════════════════════════════════
# EXCEL EXPORT — colores en formato AARRGGBB correcto
# ══════════════════════════════════════════════════════════
def exportar_excel(resultados,titulo="Scorpion Elite V4 Pro"):
    def mkfill(c): return PatternFill("solid",fgColor=f"FF{c.upper()}")
    def mkbrd():
        s=Side(border_style="thin",color="FF1A2A4A")
        return Border(left=s,right=s,top=s,bottom=s)
    def mkal(h="center"): return Alignment(horizontal=h,vertical="center",wrap_text=True)
    def sc(cell,bg="080820",fg="FFFFFF",bold=False,sz=9,h="center"):
        cell.fill=mkfill(bg)
        cell.font=Font(color=f"FF{fg.upper()}",bold=bold,size=sz,name="Arial")
        cell.alignment=mkal(h); cell.border=mkbrd()
    def hr(ws,row,vals,bg="080820",fg="FFD700",sz=9):
        for c,v in enumerate(vals,1): cl=ws.cell(row=row,column=c,value=v); sc(cl,bg,fg,True,sz)
    def tr(ws,row,text,nc,bg="080820",fg="FFD700",sz=12):
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=nc)
        sc(ws.cell(row=row,column=1,value=text),bg,fg,True,sz)
    def sw(ws,wl):
        for i,w in enumerate(wl,1): ws.column_dimensions[get_column_letter(i)].width=w

    wb=Workbook(); first=True
    COLS=["HORA","LIGA","PARTIDO","1X2","P1%","PX%","P2%","CONF",
          "xG L","xG V","xG T","O0.5","O1.5","O2.5","O3.5","BTTS",
          "CORNERS","TIROS","TARJ","PICK CLAVE","RANGO",
          "POISSON","D-COLES","M.CARLO","ELO","FUENTE"]
    WW=[9,18,30,7,7,7,7,7,7,7,7,8,8,8,8,7,14,10,12,26,7,9,9,9,9,14]
    for dia in list(dict.fromkeys(r["dia"] for r in resultados)):
        ws=wb.active if first else wb.create_sheet(title=str(dia)[:20])
        first=False
        if ws==wb.active: ws.title=str(dia)[:20]
        ws.sheet_view.showGridLines=False; sw(ws,WW)
        tr(ws,1,f"SCORPION ELITE V4 PRO | {dia}",len(COLS))
        ws.row_dimensions[1].height=24; hr(ws,2,COLS,sz=8); ws.row_dimensions[2].height=36; fi=3
        for p in sorted([x for x in resultados if x["dia"]==dia],key=lambda x:x.get("hora_sort",0)):
            rg=p.get("rango","C")
            rv=[p.get("hora",""),p.get("liga","")[:16],f'{p["local"]} vs {p["visitante"]}',
                p.get("mk","?"),f'{p.get("p1",0):.1f}%',f'{p.get("px",0):.1f}%',f'{p.get("p2",0):.1f}%',
                f'{p.get("confianza",0)}%',
                p.get("xl",0),p.get("xv",0),p.get("xt",0),
                f'{p.get("o05",0)}%',f'{p.get("over15",0)}%',f'{p.get("over25",0)}%',f'{p.get("over35",0)}%',
                f'{p.get("btts_si",0)}%',
                p.get("corners_str","")[:12],
                f'{p.get("tiros_l",0)}-{p.get("tiros_v",0)} (~{p.get("tiros_tot",0)})',
                p.get("tar_str",""),
                p.get("mk2",""),rg,
                f'{p.get("p1_po",0):.1f}%',f'{p.get("p1_dc",0):.1f}%',
                f'{p.get("p1_mc",0):.1f}%',f'{p.get("p1_el",0):.1f}%',
                f'L:{p.get("fuente_l","?")[:6]} V:{p.get("fuente_v","?")[:6]}']
            bg="0A2810" if rg=="A+" else ("0A0A28" if fi%2==0 else "0D0D1E")
            for col,val in enumerate(rv,1):
                cell=ws.cell(row=fi,column=col,value=val)
                fg="FFFFFF"
                if col==21: fg="FFD700" if rg=="A+" else ("44AAFF" if rg=="B" else "888888")
                elif col in (9,10,11): fg="00DDFF"
                elif col==26: fg="00CC44" if "API" in str(val) else "FFAA00"
                sc(cell,bg,fg,rg=="A+" and col in (3,20,21),9)
            ws.row_dimensions[fi].height=22; fi+=1

    # Hoja TOP PICKS
    wsp=wb.create_sheet(title="TOP PICKS"); wsp.sheet_view.showGridLines=False
    pc=["HORA","PARTIDO","LIGA","PICK CLAVE","P1%","PX%","P2%","xG T","O2.5","BTTS","TIROS","CORNERS","TARJ","CONF","RANGO","MARCADORES TOP"]
    sw(wsp,[9,28,16,28,7,7,7,7,8,7,10,14,12,8,7,28])
    tr(wsp,1,"TOP PICKS — Mayor confianza y ventaja estadistica",len(pc),"081208","FFD700")
    hr(wsp,2,pc,"0A1E0A","00FF88"); fp=3
    top=sorted([r for r in resultados if r.get("rango") in ("A+","B")],
               key=lambda x:(x.get("confianza",0),max(x.get("p1",0),x.get("p2",0))),reverse=True)[:15]
    for p in top:
        bg="0A2810" if p.get("rango")=="A+" else "0A0A28"
        top_mk=" | ".join([f"{k}({v}%)" for k,v in list(p.get("top_ex",{}).items())[:4]])
        rv=[p.get("hora",""),f'{p["local"]} vs {p["visitante"]}',p.get("liga","")[:14],
            p.get("mk2",""),f'{p.get("p1",0):.1f}%',f'{p.get("px",0):.1f}%',f'{p.get("p2",0):.1f}%',
            p.get("xt",0),f'{p.get("over25",0)}%',f'{p.get("btts_si",0)}%',
            f'~{p.get("tiros_tot",0)}',p.get("corners_str","")[:12],p.get("tar_str",""),
            f'{p.get("confianza",0)}%',p.get("rango",""),top_mk]
        for col,val in enumerate(rv,1):
            cell=wsp.cell(row=fp,column=col,value=val)
            fg="FFFFFF"
            if col==4: fg="FFD700"
            elif col==15: fg="00FF88" if p.get("rango")=="A+" else "44AAFF"
            elif col==16: fg="888888"
            sc(cell,bg,fg,col in (4,15),9)
        wsp.row_dimensions[fp].height=20; fp+=1

    # Hoja ESCALERA
    wse=wb.create_sheet(title="ESCALERA"); wse.sheet_view.showGridLines=False
    ec=["PASO","HORA","PARTIDO","LIGA","PICK","PROB%","C.JUSTA","ESTADO"]
    sw(wse,[7,9,28,18,26,8,9,14])
    tr(wse,1,"RETO ESCALERA — Picks Seleccionados del Dia",len(ec),"081208","FFD700")
    hr(wse,2,ec,"0A1E0A","FFD700"); fe=3
    esc=sorted([r for r in resultados if r.get("rango") in ("A+","B")],
               key=lambda x:(x.get("confianza",0)),reverse=True)[:12]
    for i,p in enumerate(esc,1):
        bg="0A2810" if i%2==0 else "0D1E10"
        cj=round(100/max(p.get("p1",50),1),2)
        rv=[i,p.get("hora",""),f'{p["local"]} vs {p["visitante"]}',
            p.get("liga","")[:16],p.get("mk2",""),
            f'{max(p.get("p1",0),p.get("p2",0)):.1f}%',str(cj),"🟡 PENDIENTE"]
        for col,val in enumerate(rv,1):
            cell=wse.cell(row=fe,column=col,value=val)
            fg="FFD700" if col==1 else ("00FF88" if col==8 else "FFFFFF")
            sc(cell,bg,fg,col==1,9)
        wse.row_dimensions[fe].height=20; fe+=1

    for ws_tab in wb.worksheets: ws_tab.freeze_panes="A3"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

# ══════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════
st.set_page_config(page_title="Scorpion Elite V4",page_icon="🦂",layout="wide")
st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif}
.stApp{background:#07070e}
.hdr{background:linear-gradient(135deg,#08102a,#10102e,#181838);
  border:1px solid #cc9900;border-radius:16px;padding:1.5rem 2rem;
  text-align:center;margin-bottom:1.5rem}
.hdr h1{color:#ffd700;font-size:2.2rem;margin:0;letter-spacing:4px;text-shadow:0 0 30px #ffd70055}
.hdr p{color:#888;margin:.3rem 0 0;font-size:.88rem}
.lbox{background:#0e0e22;border:1px solid #cc9900;border-radius:16px;padding:2.5rem;max-width:440px;margin:1.5rem auto}
.mc{background:#0e0e22;border:1px solid #1a2a44;border-radius:10px;padding:1rem;text-align:center;margin:.3rem 0}
.mc .v{color:#ffd700;font-size:1.8rem;font-weight:700}
.mc .l{color:#555;font-size:.7rem;margin-top:3px;text-transform:uppercase;letter-spacing:1px}
.badge-AP{background:#082a12;color:#00ee66;border:1px solid #00aa44;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:700}
.badge-B{background:#081840;color:#44aaff;border:1px solid #2255aa;padding:2px 9px;border-radius:20px;font-size:11px}
.badge-C{background:#1a1200;color:#ccaa00;border:1px solid #886600;padding:2px 9px;border-radius:20px;font-size:11px}
.badge-D{background:#200a0a;color:#ee4444;border:1px solid #882222;padding:2px 9px;border-radius:20px;font-size:11px}
.pill-ok{background:#082a12;color:#00ee66;padding:2px 10px;border-radius:20px;font-size:11px}
.pill-v{background:#2a0808;color:#ee4444;padding:2px 10px;border-radius:20px;font-size:11px}
.pill-f{background:#1a1500;color:#ccaa00;padding:2px 10px;border-radius:20px;font-size:11px}
.pick-box{background:#0e0e22;border:1px solid #1a2a44;border-radius:10px;padding:.9rem 1rem;margin:.4rem 0}
.pick-box.ap{border-color:#00aa44;background:#081a10}
.pick-box.b{border-color:#2255aa;background:#08101a}
.mkt-row{background:#0a0a1e;border-radius:8px;padding:.5rem .9rem;margin:.2rem 0;
  display:flex;justify-content:space-between;align-items:center}
.mkt-name{color:#cccccc;font-size:.88rem}
.mkt-prob{color:#ffd700;font-weight:700;font-size:.95rem}
.mkt-edge-pos{color:#00ee66;font-size:.8rem}
.mkt-edge-neg{color:#ee4444;font-size:.8rem}
.pub-btn{background:#ffd700;color:#000;border:none;border-radius:6px;padding:4px 14px;
  font-size:.8rem;font-weight:700;cursor:pointer}
.esc-box{background:#081a10;border:1px solid #00aa44;border-radius:10px;padding:.8rem 1rem;margin:.3rem 0}
.comb-box{background:#0a101a;border:1px solid #2255aa;border-radius:10px;padding:.8rem 1rem;margin:.3rem 0}
.ft{text-align:center;color:#1e1e2e;font-size:.68rem;margin-top:2rem;padding-top:.8rem;border-top:1px solid #10102a}
div[data-testid="stButton"] button{background:linear-gradient(135deg,#ffd700,#cc7700)!important;
  color:#000!important;font-weight:700!important;border:none!important;border-radius:8px!important;width:100%}
</style>""",unsafe_allow_html=True)

def hdr(sub="Motor de Analisis Deportivo — 4 Modelos + Datos Reales"):
    st.markdown(f'<div class="hdr"><h1>🦂 SCORPION ELITE</h1><p>{sub}</p></div>',unsafe_allow_html=True)

def bdg(r):
    return f'<span class="badge-{r}">{r}</span>'

def pll(plan,dr):
    if plan=="gratis": return '<span class="pill-f">Plan Gratis</span>'
    if plan in ("dia","semana","mes","admin"):
        return f'<span class="pill-ok">{"Admin" if plan=="admin" else plan.upper()} · {dr} dias</span>'
    return '<span class="pill-v">Vencido</span>'

def widget_archivo(max_p=None,key="up"):
    up=st.file_uploader("Sube Excel o captura de pantalla",type=["xlsx","xls","png","jpg","jpeg","webp"],key=key)
    if not up: return None
    fb=up.read(); nm=up.name.lower()
    if nm.endswith((".xlsx",".xls")):
        try: p=leer_excel(fb)
        except Exception as e: st.error(f"Error: {e}"); return None
    else:
        mt="image/png" if nm.endswith(".png") else ("image/webp" if nm.endswith(".webp") else "image/jpeg")
        with st.spinner("Leyendo imagen con IA..."): p=leer_imagen(fb,mt)
    if not p: st.warning("No se encontraron partidos."); return None
    if max_p: p=p[:max_p]
    st.success(f"✅ {len(p)} partidos detectados"); return p

def mostrar_mercados_con_publicar(calc, partido):
    """Muestra mercados organizados por categoria con botones de publicar."""
    mercados=calc.get("mercados_picks",[])
    if not mercados: return
    picks_sel=st.session_state.get("picks_sel",[])

    # Organizar por categorias
    cats = {
        "⚽ Resultado 1X2":   [],
        "🎯 Goles Over/Under":[],
        "🤝 BTTS":            [],
        "📐 Corners":         [],
        "🎽 Tarjetas":        [],
        "🔫 Tiros":           [],
    }
    for item in mercados[:16]:
        nombre,prob,cuota = item
        n=nombre.lower()
        if any(x in n for x in ["victoria","empate","local","visita","(1)","(2)","(x)"]):
            cats["⚽ Resultado 1X2"].append(item)
        elif "btts" in n or "ambos" in n:
            cats["🤝 BTTS"].append(item)
        elif "corner" in n or "esquina" in n:
            cats["📐 Corners"].append(item)
        elif "tarjet" in n or "amarill" in n:
            cats["🎽 Tarjetas"].append(item)
        elif "tiro" in n or "shot" in n:
            cats["🔫 Tiros"].append(item)
        elif "over" in n or "under" in n or "gol" in n:
            cats["🎯 Goles Over/Under"].append(item)

    st.markdown("""
    <style>
    .mkt-cat{color:#ffd700;font-weight:600;font-size:.85rem;margin:12px 0 4px;
      border-bottom:1px solid #ffd70033;padding-bottom:4px}
    .mkt-row2{display:grid;grid-template-columns:3fr 1fr 1.2fr 1.2fr 1.5fr;
      align-items:center;background:#0a0a1e;border-radius:8px;
      padding:7px 10px;margin:3px 0;gap:8px}
    .mkt-n{color:#ddd;font-size:.85rem}
    .mkt-p{color:#ffd700;font-weight:700;font-size:.95rem;text-align:center}
    .mkt-c{color:#888;font-size:.78rem;text-align:center}
    .mkt-ep{color:#00ee66;font-size:.8rem;font-weight:600;text-align:center}
    .mkt-en{color:#ff4444;font-size:.8rem;font-weight:600;text-align:center}
    .val-badge{background:#082a12;color:#00ee66;border:1px solid #00aa44;
      padding:1px 6px;border-radius:4px;font-size:.72rem;font-weight:700}
    .neutro-badge{background:#1a1200;color:#ccaa00;padding:1px 6px;border-radius:4px;font-size:.72rem}
    .neg-badge{background:#200a0a;color:#ff4444;padding:1px 6px;border-radius:4px;font-size:.72rem}
    </style>
    """, unsafe_allow_html=True)

    # Cabecera de columnas
    hc1,hc2,hc3,hc4,hc5=st.columns([3,1,1.2,1.2,1.5])
    with hc1: st.markdown('<span style="color:#888;font-size:.75rem">MERCADO</span>',unsafe_allow_html=True)
    with hc2: st.markdown('<span style="color:#888;font-size:.75rem">PROB</span>',unsafe_allow_html=True)
    with hc3: st.markdown('<span style="color:#888;font-size:.75rem">C.REF</span>',unsafe_allow_html=True)
    with hc4: st.markdown('<span style="color:#888;font-size:.75rem">EDGE</span>',unsafe_allow_html=True)
    with hc5: st.markdown('<span style="color:#888;font-size:.75rem">ACCION</span>',unsafe_allow_html=True)

    idx_global=0
    hay_valor=False
    for cat_name, items in cats.items():
        # Solo mostrar mercados con edge >= 0
        # Edge real: solo mostrar si la cuota de mercado es MEJOR que la justa
        # Con cuotas de referencia propias, no hay edge real sin odds reales
        # Mostramos todos los mercados pero marcamos si son valor o estimado
        items_valor=[(n,p,c) for n,p,c in items if c and p>0]
        if not items_valor: continue
        hay_valor=True
        st.markdown(f'<div class="mkt-cat">{cat_name}</div>', unsafe_allow_html=True)
        for nombre,prob,cuota in items_valor:
            edge=round((prob/100*cuota-1)*100,1) if cuota else 0
            sel=any(p.get("mercado")==nombre and p.get("local")==partido["local"] for p in picks_sel)
            # Badge de valor
            # Con cuotas de referencia propias, edge siempre es ~0
            # Solo hay edge real si se configura ODDS_API_KEY con cuotas en vivo
            tiene_odds_reales = False  # TODO: activar cuando ODDS_API_KEY este configurada
            if tiene_odds_reales and edge>=5:   badge='<span class="val-badge">🔥 VALOR REAL</span>'
            elif tiene_odds_reales and edge>=2: badge='<span class="val-badge">✅ VALOR REAL</span>'
            else:
                # Mostrar probabilidad del modelo como referencia
                if prob>=70:   badge='<span class="val-badge">⭐ Alta prob</span>'
                elif prob>=55: badge='<span class="neutro-badge">📊 Buena prob</span>'
                else:          badge='<span class="neg-badge">📉 Prob baja</span>'
            edge_cls="mkt-ep" if edge>=0 else "mkt-en"
            c1,c2,c3,c4,c5=st.columns([3,1,1.2,1.2,1.5])
            with c1: st.markdown(f'{badge} <span class="mkt-n">{nombre}</span>',unsafe_allow_html=True)
            with c2: st.markdown(f'<span class="mkt-p">{prob:.0f}%</span>',unsafe_allow_html=True)
            with c3: st.markdown(f'<span class="mkt-c">{cuota}</span>',unsafe_allow_html=True)
            with c4: st.markdown(f'<span class="{edge_cls}">{edge:+.1f}%</span>',unsafe_allow_html=True)
            with c5:
                key_btn=f"pub_{partido['local'][:6]}_{partido['visitante'][:6]}_{idx_global}"
                btn_lbl="✔ Agregado" if sel else "➕ Seleccionar"
                if st.button(btn_lbl,key=key_btn):
                    pick_data={"local":partido["local"],"visitante":partido["visitante"],
                               "liga":partido.get("liga",""),"hora":partido.get("hora","00:00"),
                               "mercado":nombre,"prob":prob,"cuota":cuota,"edge":edge,
                               "confianza":calc.get("confianza",0),"rango":calc.get("rango","C")}
                    if not sel:
                        if "picks_sel" not in st.session_state: st.session_state.picks_sel=[]
                        st.session_state.picks_sel.append(pick_data)
                    else:
                        st.session_state.picks_sel=[p for p in st.session_state.picks_sel
                            if not(p.get("mercado")==nombre and p.get("local")==partido["local"])]
                    st.rerun()
            idx_global+=1

# ══════════════════════════════════════════════════════════
# PANTALLA LOGIN
# ══════════════════════════════════════════════════════════
def pantalla_login():
    hdr()
    st.markdown('<div class="lbox">',unsafe_allow_html=True)
    st.markdown("### 🔐 Acceder")
    ced=st.text_input("Cedula / DNI",placeholder="Tu cedula o 'admin'")
    es_adm=ced.strip().lower()=="admin"
    pwd=st.text_input("Contrasena admin",type="password") if es_adm else ""
    if st.button("Entrar →"):
        if not ced.strip(): st.error("Ingresa tu cedula."); return
        if es_adm:
            if db_login_admin(pwd.strip()):
                u=db_get("admin")
                st.session_state.update({"li":True,"u":u,"ced":ced.strip(),"picks_sel":[]}); st.rerun()
            else: st.error("Contrasena incorrecta.")
        else:
            u=db_get(ced.strip())
            if not u:
                db_guardar_usuario(ced.strip(),f"Usuario {ced.strip()[:6]}","gratis",36500,date.today())
                u=db_get(ced.strip())
            ok,plan,dr=db_acceso(ced.strip())
            if not ok: st.error("Acceso vencido o inactivo. Contacta al administrador.")
            else: st.session_state.update({"li":True,"u":u,"ced":ced.strip()}); st.rerun()
    st.markdown('</div>',unsafe_allow_html=True)
    st.markdown("---")
    c1,c2,c3,c4=st.columns(4)
    infos=[("🆓 Gratis","Sube Excel/imagen\nMax 5 partidos/dia\nPicks del dia (limitados)"),
           ("📅 Plan Dia","Liga a eleccion\nPartidos del dia\nExcel con 4 modelos"),
           ("📆 Plan Semana","Semana completa\nDias especificos\nMulti-liga + Escalera"),
           ("👑 Plan Mes","Todo ilimitado\nTodas las ligas\nHistorial + Combinadas")]
    for col,(t,d) in zip([c1,c2,c3,c4],infos):
        with col: st.info(f"**{t}**\n\n{d}")
    st.markdown('<div class="ft">🦂 Scorpion Elite V4 Pro 2025 · Solo uso informativo · Las apuestas implican riesgo real de perdida</div>',unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# PANTALLA ADMIN
# ══════════════════════════════════════════════════════════
def pantalla_admin():
    hdr("Panel de Administrador")
    s=db_stats()
    c1,c2,c3,c4=st.columns(4)
    for col,(v,l) in zip([c1,c2,c3,c4],[(s["usuarios"],"Usuarios"),(s["activos"],"Activos"),(s["picks"],"Picks pub."),(s["historial"],"Analisis")]):
        with col: st.markdown(f'<div class="mc"><div class="v">{v}</div><div class="l">{l}</div></div>',unsafe_allow_html=True)
    st.markdown("---")

    tabs=st.tabs(["🔍 Analizar Partido","🏟️ Analizar por Liga","📢 Publicar Picks","👥 Clientes","📊 Escalera"])

    # ── Tab 1: Analizar partido individual ─────────────────
    with tabs[0]:
        st.markdown("### Analizar cualquier partido")
        st.info("Escribe cualquier partido: Millonarios vs Nacional, Barcelona vs Real Madrid, etc.")
        c1,c2=st.columns(2)
        with c1:
            local_i  =st.text_input("Equipo Local",placeholder="ej: Millonarios")
            visita_i =st.text_input("Equipo Visitante",placeholder="ej: Atletico Nacional")
            liga_i   =st.text_input("Liga / Torneo",placeholder="ej: Liga BetPlay, Premier League")
            hora_i   =st.text_input("Hora (opcional)","20:00")
        with c2:
            fecha_i  =st.date_input("Fecha",value=date.today())
            st.markdown("**Usar datos reales de API**")
            usar_api_i=st.checkbox("Buscar stats reales en internet",value=True)
        if st.button("🦂 Analizar este partido",key="btn_individual"):
            if local_i and visita_i and liga_i:
                partido_i={"dia":str(fecha_i),"hora":hora_i,"hora_sort":0,
                           "liga":liga_i,"liga_id":0,"local":local_i,"visitante":visita_i,
                           "tid_l":None,"tid_v":None}
                with st.spinner(f"Analizando {local_i} vs {visita_i}..."):
                    if usar_api_i:
                        # Buscar liga_id automaticamente si el nombre coincide
                        liga_i_lower = liga_i.lower()
                        liga_id_auto = next((v for k,v in LIGAS.items() if
                                            any(w in liga_i_lower for w in k.lower().split()
                                                if len(w)>3)), None)
                        sl_det=obtener_stats_detalle(local_i, liga_i, lid=liga_id_auto)
                        sv_det=obtener_stats_detalle(visita_i, liga_i, lid=liga_id_auto)
                        gml=sl_det["gm"] if sl_det["ok"] else None; gcl=sl_det["gc"] if sl_det["ok"] else None
                        gmv=sv_det["gm"] if sv_det["ok"] else None; gcv=sv_det["gc"] if sv_det["ok"] else None
                        elo_l=sl_det["elo"]; elo_v=sv_det["elo"]
                        fl=sl_det["fuente"]; fv=sv_det["fuente"]
                    else:
                        sl_det={}; sv_det={}
                        gml=gcl=gmv=gcv=elo_l=elo_v=None; fl=fv="Manual"
                    calc=calcular(gml,gcl,gmv,gcv,liga_i,elo_l,elo_v)
                resultado_i={**partido_i,**calc,"fuente_l":fl,"fuente_v":fv}
                st.session_state["resultado_individual"]=resultado_i
                st.session_state["sl_det"]=sl_det
                st.session_state["sv_det"]=sv_det
                fuentes_ok = sl_det.get("ok",False) and sv_det.get("ok",False)
                if fuentes_ok:
                    st.success(f"✅ Analisis con DATOS REALES — Rango {calc['rango']} ({calc['confianza']}% confianza) | Fuente: {sl_det.get('fuente','')} / {sv_det.get('fuente','')}")
                else:
                    st.warning(f"⚠️ Rango {calc['rango']} ({calc['confianza']}% confianza) — Sin datos reales encontrados. Se usaron promedios de liga. Los valores de tiros/corners/tarjetas son ESTIMADOS, no reales.")
            else:
                st.error("Completa local, visitante y liga.")

        if "resultado_individual" in st.session_state:
            r = st.session_state["resultado_individual"]
            sl_det = st.session_state.get("sl_det", {})
            sv_det = st.session_state.get("sv_det", {})
            st.markdown("---")
            st.markdown(f"## {bdg(r['rango'])} **{r['local']} vs {r['visitante']}** — {r['liga']}", unsafe_allow_html=True)

            # ── SECCION 1: Comparacion de modelos (arriba, ancho completo) ──
            st.markdown("### 📊 Comparacion de modelos")
            mc1,mc2,mc3,mc4,mc5,mc6 = st.columns(6)
            with mc1: st.markdown(f'<div class="mc"><div class="v">{r["p1"]:.0f}%</div><div class="l">P1 Final</div></div>',unsafe_allow_html=True)
            with mc2: st.markdown(f'<div class="mc"><div class="v">{r["px"]:.0f}%</div><div class="l">Empate</div></div>',unsafe_allow_html=True)
            with mc3: st.markdown(f'<div class="mc"><div class="v">{r["p2"]:.0f}%</div><div class="l">P2 Final</div></div>',unsafe_allow_html=True)
            with mc4: st.markdown(f'<div class="mc"><div class="v">{r["confianza"]}%</div><div class="l">Confianza</div></div>',unsafe_allow_html=True)
            with mc5: st.markdown(f'<div class="mc"><div class="v">{r["xt"]}</div><div class="l">xG Total</div></div>',unsafe_allow_html=True)
            with mc6: st.markdown(f'<div class="mc"><div class="v">{r["rango"]}</div><div class="l">Rango</div></div>',unsafe_allow_html=True)

            # Tabla modelos HTML pura
            lbl_l = r["local"][:14]; lbl_v = r["visitante"][:14]
            filas_m = [
                ("Poisson",       r["p1_po"], r["px_po"], r["p2_po"]),
                ("Dixon-Coles",   r["p1_dc"], r["px_dc"], r["p2_dc"]),
                ("Monte Carlo",   r["p1_mc"], r["px_mc"], r["p2_mc"]),
                ("Elo",           r["p1_el"], r["px_el"], r["p2_el"]),
                ("⭐ FINAL",      r["p1"],    r["px"],    r["p2"]),
            ]
            tbl = f'<table style="width:100%;border-collapse:collapse;font-size:.84rem;margin:8px 0"><thead><tr style="background:#0a0a1e;color:#ffd700"><th style="padding:8px 12px;text-align:left">Modelo</th><th style="padding:8px 12px;text-align:center">{lbl_l} (P1)%</th><th style="padding:8px 12px;text-align:center">Empate%</th><th style="padding:8px 12px;text-align:center">{lbl_v} (P2)%</th></tr></thead><tbody>'
            for i,(mod,p1v,pxv,p2v) in enumerate(filas_m):
                bg="#0d1e0d" if mod.startswith("⭐") else ("#0a0a1e" if i%2==0 else "#0d0d18")
                fw="700" if mod.startswith("⭐") else "400"
                clr="#ffd700" if mod.startswith("⭐") else "#fff"
                mx=max(p1v,pxv,p2v)
                def c(v,mx=mx,clr=clr): return f'<td style="padding:7px 12px;text-align:center;{"color:#00ee66;font-weight:700" if v==mx else f"color:{clr}"}">{ v:.1f}%</td>'
                tbl+=f'<tr style="background:{bg};font-weight:{fw}"><td style="padding:7px 12px;color:{clr}">{mod}</td>{c(p1v)}{c(pxv)}{c(p2v)}</tr>'
            tbl+="</tbody></table>"
            st.markdown(tbl, unsafe_allow_html=True)

            # Stats del partido en tabla organizada
            st.markdown("**📊 Estadisticas del partido**")
            local_n = r["local"][:16]; visita_n = r["visitante"][:16]
            tbl_stats = f"""<table style="width:100%;border-collapse:collapse;font-size:.83rem;margin:6px 0">
              <thead><tr style="background:#0a0a1e;color:#ffd700">
                <th style="padding:6px 12px;text-align:left">Metrica</th>
                <th style="padding:6px 12px;text-align:center">{local_n}</th>
                <th style="padding:6px 12px;text-align:center">Total</th>
                <th style="padding:6px 12px;text-align:center">{visita_n}</th>
              </tr></thead><tbody>
              <tr style="background:#0d0d18">
                <td style="padding:5px 12px;color:#aaa">xG Esperado</td>
                <td style="padding:5px 12px;text-align:center;color:#00ddff;font-weight:700">{r.get("xl",0)}</td>
                <td style="padding:5px 12px;text-align:center;color:#888">{r.get("xt",0)}</td>
                <td style="padding:5px 12px;text-align:center;color:#00ddff;font-weight:700">{r.get("xv",0)}</td>
              </tr>
              <tr style="background:#0a0a1e">
                <td style="padding:5px 12px;color:#aaa">Tiros al arco (est.)</td>
                <td style="padding:5px 12px;text-align:center;color:#fff">~{r.get("tiros_l",0)}</td>
                <td style="padding:5px 12px;text-align:center;color:#888">~{r.get("tiros_tot",0)}</td>
                <td style="padding:5px 12px;text-align:center;color:#fff">~{r.get("tiros_v",0)}</td>
              </tr>
              <tr style="background:#0d0d18">
                <td style="padding:5px 12px;color:#aaa">Corners (est.)</td>
                <td style="padding:5px 12px;text-align:center;color:#fff" colspan="3">{r.get("corners_str","")}</td>
              </tr>
              <tr style="background:#0a0a1e">
                <td style="padding:5px 12px;color:#aaa">Tarjetas (est.)</td>
                <td style="padding:5px 12px;text-align:center;color:#fff" colspan="3">{r.get("tar_str","")}</td>
              </tr>
              <tr style="background:#0d0d18">
                <td style="padding:5px 12px;color:#aaa">Over 1.5 / 2.5 / 3.5</td>
                <td style="padding:5px 12px;text-align:center;color:#ffd700;font-weight:700" colspan="3">{r.get("over15",0)}% / {r.get("over25",0)}% / {r.get("over35",0)}%</td>
              </tr>
              <tr style="background:#0a0a1e">
                <td style="padding:5px 12px;color:#aaa">BTTS (Ambos Marcan)</td>
                <td style="padding:5px 12px;text-align:center;color:#ffd700;font-weight:700" colspan="3">{r.get("btts_si",0)}% Si / {r.get("btts_no",0)}% No</td>
              </tr>
            </tbody></table>"""
            st.markdown(tbl_stats, unsafe_allow_html=True)
            # Marcadores mas probables
            top_ex_str = " | ".join([f"{k} ({v}%)" for k,v in list(r.get("top_ex",{}).items())[:5]])
            st.markdown(f"🎯 **Marcadores mas probables:** {top_ex_str}")
            col_a, col_b = st.columns([1,1])
            with col_a: pass

            # ── SECCION 2: Tabla de datos reales de cada equipo ─────────────
            st.markdown("---")
            st.markdown("### 🌐 Datos reales encontrados en internet")
            if sl_det or sv_det:
                col_l, col_v = st.columns(2)
                for col, det, nombre_eq in [(col_l, sl_det, r["local"]), (col_v, sv_det, r["visitante"])]:
                    with col:
                        fuente_ico = "✅" if det.get("ok") else "⚠️"
                        fuente_txt = det.get("fuente","Sin datos")
                        st.markdown(f"**{fuente_ico} {nombre_eq}**")
                        st.caption(f"Fuente: {fuente_txt}")
                        fuentes_str = " + ".join(det.get("fuentes_usadas",[])) or "Sin datos"
                        st.caption(f"Fuentes consultadas: {fuentes_str}")
                        if det.get("ok"):
                            gm   = det.get("gm");   gc   = det.get("gc")
                            elo  = det.get("elo");   xg   = det.get("xg")
                            tiros= det.get("tiros_pg"); tarj=det.get("tarj_pg")
                            u5   = det.get("ultimos5",[])
                            g5   = det.get("ganados5",0); e5=det.get("empatados5",0); p5=det.get("perdidos5",0)
                            gf5  = det.get("goles_fav5",0); gc5=det.get("goles_con5",0)

                            # Tabla estilo Sofascore
                            rows = [
                                ("⚽ Goles marcados/partido",   f"{gm:.2f}" if gm else "N/D",   "#00ee66"),
                                ("🥅 Goles concedidos/partido", f"{gc:.2f}" if gc else "N/D",   "#ff6666"),
                                ("🎯 xG por partido",           f"{xg:.2f}" if xg else "N/D",   "#00ddff"),
                                ("🔫 Tiros al arco/partido",    f"{tiros:.1f}" if tiros else "N/D","#ffaa00"),
                                ("🟨 Tarjetas amarillas/partido",f"{tarj:.2f}" if tarj else "N/D","#ffd700"),
                                ("🏆 ELO Rating",               f"{int(elo)}" if elo else "N/D", "#ffd700"),
                                ("📊 Ultimos 5 (G/E/P)",        f"{g5}G / {e5}E / {p5}P",       "#ffffff"),
                                ("⚽ Goles U5",                 f"{gf5} a favor / {gc5} en contra","#aaaaaa"),
                            ]
                            tbl_eq = '<table style="width:100%;border-collapse:collapse;font-size:.82rem;margin:6px 0">'
                            tbl_eq += '<thead><tr style="background:#0a0a1e;color:#ffd700"><th style="padding:6px 10px;text-align:left">Estadistica</th><th style="padding:6px 10px;text-align:center">Valor</th></tr></thead><tbody>'
                            for i_r,(stat,val,clr) in enumerate(rows):
                                bg_r="#0d0d18" if i_r%2==0 else "#0a0a1e"
                                tbl_eq+=f'<tr style="background:{bg_r}"><td style="padding:5px 10px;color:#aaa">{stat}</td><td style="padding:5px 10px;text-align:center;color:{clr};font-weight:700">{val}</td></tr>'
                            tbl_eq+="</tbody></table>"
                            st.markdown(tbl_eq, unsafe_allow_html=True)

                            # Ultimos 5 partidos detallados
                            if u5:
                                st.markdown("**🕐 Ultimos 5 partidos:**")
                                tbl_u5='<table style="width:100%;border-collapse:collapse;font-size:.8rem;margin:4px 0">'
                                tbl_u5+='<thead><tr style="background:#0a0a1e;color:#ffd700"><th style="padding:5px 8px;text-align:left">Rival</th><th style="padding:5px 8px;text-align:center">Cond.</th><th style="padding:5px 8px;text-align:center">Marcador</th><th style="padding:5px 8px;text-align:center">Res</th></tr></thead><tbody>'
                                for i_u,pu in enumerate(u5):
                                    bg_u="#0d0d18" if i_u%2==0 else "#0a0a1e"
                                    rc={"G":"#00ee66","E":"#ffaa00","P":"#ff4444"}.get(pu["res"],"#fff")
                                    tbl_u5+=f'<tr style="background:{bg_u}"><td style="padding:5px 8px;color:#ccc">{pu["rival"]}</td><td style="padding:5px 8px;text-align:center;color:#888;font-size:.75rem">{pu["local"]}</td><td style="padding:5px 8px;text-align:center;color:#fff;font-weight:700">{pu["goles"]}</td><td style="padding:5px 8px;text-align:center"><span style="background:{rc};color:#000;padding:1px 8px;border-radius:4px;font-weight:700;font-size:.8rem">{pu["res"]}</span></td></tr>'
                                tbl_u5+="</tbody></table>"
                                st.markdown(tbl_u5, unsafe_allow_html=True)
                            else:
                                st.caption("Historial partido a partido no disponible en esta fuente.")
                        else:
                            st.warning(f"Sin datos reales para {nombre_eq}. Se usaron promedios de la liga.")
                            st.caption("Verifica que la API_FOOTBALL_KEY este configurada en Streamlit Secrets.")
            else:
                st.info("Activa 'Buscar stats reales en internet' para ver datos de los equipos.")

            # ── SECCION 3: Solo picks con VALOR (edge positivo) ─────────────
            st.markdown("---")
            st.markdown("### 📌 Picks con valor estadistico — selecciona para publicar")
            st.caption("⚠️ Las probabilidades son del modelo matematico. El EDGE real solo se calcula cuando configuras ODDS_API_KEY en Secrets con cuotas de tu casa de apuestas. Compara la cuota justa con la cuota real que ofrece tu casa antes de apostar.")
            mostrar_mercados_con_publicar(r, r)

            # ── SECCION 4: Picks seleccionados ──────────────────────────────
            picks_sel = st.session_state.get("picks_sel", [])
            if picks_sel:
                st.markdown(f"---\n### ✅ Picks seleccionados ({len(picks_sel)})")
                plan_pub = st.selectbox("Plan minimo", ["gratis","dia","semana","mes"], key="plan_pub_i")
                notas_pub = st.text_area("Notas (opcional)", height=60, key="notas_pub_i")
                if st.button("📢 PUBLICAR PICKS SELECCIONADOS", key="pub_sel"):
                    for pk in picks_sel:
                        det_str = f"xG:{r.get('xl',0)}-{r.get('xv',0)} | O2.5:{r.get('over25',0)}% | BTTS:{r.get('btts_si',0)}% | Tiros:~{r.get('tiros_tot',0)} | Corners:{r.get('corners_str','')}"
                        db_pick_guardar(str(r.get("dia",date.today())),pk["liga"],pk["local"],pk["visitante"],
                                        pk["hora"],pk["mercado"],det_str,pk["cuota"],pk["edge"],
                                        pk["confianza"],pk["rango"],notas_pub,plan_pub,0)
                    st.success(f"✅ {len(picks_sel)} picks publicados!")
                    st.session_state.picks_sel = []; st.rerun()
                if st.button("🗑️ Limpiar seleccion", key="clear_sel"):
                    st.session_state.picks_sel = []; st.rerun()
                for pk in picks_sel:
                    st.markdown(f"✔ **{pk['local']} vs {pk['visitante']}** · {pk['mercado']} · {pk['prob']:.0f}% · Edge:{pk['edge']:+.1f}%")

    # ── Tab 2: Analizar por liga ────────────────────────────
    with tabs[1]:
        st.markdown("### Analizar liga completa")
        c1,c2=st.columns([2,1])
        with c1:
            ligas_sel=st.multiselect("Ligas",list(LIGAS.keys()),default=[list(LIGAS.keys())[0]])
        with c2:
            ml=st.radio("Periodo",["Hoy","Dia especifico","Esta semana","Semana personalizada"])
            if ml=="Hoy": fecha_s=date.today(); modo="dia"
            elif ml=="Dia especifico": fecha_s=st.date_input("Fecha",key="fa2"); modo="dia"
            elif ml=="Esta semana":
                hoy=date.today(); lu=hoy-timedelta(days=hoy.weekday())
                fd=lu; fh=lu+timedelta(days=6); modo="rango"
            else:
                fd=st.date_input("Desde",value=date.today(),key="fd2")
                fh=st.date_input("Hasta",value=date.today()+timedelta(days=6),key="fh2"); modo="rango"

        if st.button("🦂 Obtener y Analizar",key="btn_liga_admin"):
            todos=[]
            with st.spinner("Obteniendo fixtures..."):
                for ln in ligas_sel:
                    lid=LIGAS[ln]
                    if modo=="dia": fx=get_fx_dia(lid,str(fecha_s))
                    else: fx=get_fx_rango(lid,str(fd),str(fh))
                    todos+=[fx2p(f) for f in fx]; time.sleep(0.3)
            if not todos:
                st.warning("No se encontraron partidos.")
            else:
                prg=st.progress(0,"Analizando...")
                res=analizar_lista(todos,usar_api=True,prog=prg)
                prg.progress(1.0,"Listo ✅")
                st.session_state["res_liga_admin"]=res
                ap=sum(1 for r in res if r["rango"]=="A+")
                c1,c2,c3=st.columns(3)
                with c1: st.markdown(f'<div class="mc"><div class="v">{len(res)}</div><div class="l">Partidos</div></div>',unsafe_allow_html=True)
                with c2: st.markdown(f'<div class="mc"><div class="v">{ap}</div><div class="l">Rango A+</div></div>',unsafe_allow_html=True)
                with c3:
                    cp=round(sum(r.get("confianza",0) for r in res)/len(res)) if res else 0
                    st.markdown(f'<div class="mc"><div class="v">{cp}%</div><div class="l">Conf. prom</div></div>',unsafe_allow_html=True)

        if "res_liga_admin" in st.session_state:
            res=st.session_state["res_liga_admin"]
            top=sorted([r for r in res if r.get("rango") in ("A+","B")],
                       key=lambda x:(x.get("confianza",0),max(x.get("p1",0),x.get("p2",0))),reverse=True)
            st.markdown("### Selecciona partidos para publicar picks automaticos")
            auto_top=st.slider("Publicar top N picks automaticamente",1,min(10,len(top)),3)
            plan_auto=st.selectbox("Plan minimo",["gratis","dia","semana","mes"],key="plan_auto")
            if st.button("🤖 Publicar top picks automaticamente"):
                count=0
                for r in top[:auto_top]:
                    mks=r.get("mercados_picks",[])
                    if mks:
                        nom,prob,cuota=mks[0]
                        edge=round((prob/100*cuota-1)*100,1) if cuota else 0
                        det=f"xG:{r.get('xl',0)}-{r.get('xv',0)} | O2.5:{r.get('over25',0)}% | Conf:{r.get('confianza',0)}%"
                        db_pick_guardar(r.get("dia",str(date.today())),r.get("liga",""),
                                        r["local"],r["visitante"],r.get("hora",""),
                                        nom,det,cuota,edge,r.get("confianza",0),
                                        r.get("rango","C"),"Auto-publicado",plan_auto,1)
                        count+=1
                st.success(f"✅ {count} picks auto-publicados!")
            xl=exportar_excel(res,"Scorpion Elite V4 — Admin")
            fl=str(fecha_s) if modo=="dia" else f"{fd}_al_{fh}"
            st.download_button("⬇️ Descargar Excel",data=xl,
                file_name=f"Scorpion_admin_{fl}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── Tab 3: Picks publicados ─────────────────────────────
    with tabs[2]:
        st.markdown("### Picks publicados")
        picks=db_picks_get(plan="admin")
        hoy_str=str(date.today())
        hoy_picks=[p for p in picks if p.get("fecha")==hoy_str]
        st.markdown(f"**Hoy ({hoy_str}): {len(hoy_picks)} picks publicados**")
        for p in hoy_picks:
            cls="ap" if p.get("rango")=="A+" else "b"
            auto_tag=" 🤖" if p.get("auto") else ""
            st.markdown(f'<div class="pick-box {cls}"><b>{p["liga"]}</b> · {p["local"]} vs {p["visitante"]} · {p.get("hora","")}{auto_tag}<br>📌 <b>{p["mercado"]}</b> @ {p.get("cuota","?")} · Edge:{p.get("edge","?")}% · Conf:{p.get("confianza","?")}% · Plan min: <b>{p.get("plan_min","gratis").upper()}</b><br><small>{p.get("detalle","")}</small></div>',unsafe_allow_html=True)

    # ── Tab 4: Clientes ─────────────────────────────────────
    with tabs[3]:
        c1,c2=st.columns(2)
        with c1:
            st.markdown("### ➕ Registrar / Actualizar")
            ced2=st.text_input("Cedula"); nom2=st.text_input("Nombre")
            plan2=st.selectbox("Plan",["gratis","dia","semana","mes"])
            dm={"gratis":30,"dia":1,"semana":7,"mes":30}
            dias2=st.number_input("Dias",min_value=1,max_value=3650,value=int(dm[plan2]))
            fi2=st.date_input("Inicio",value=date.today())
            if st.button("💾 Guardar"):
                if ced2.strip():
                    db_guardar_usuario(ced2.strip(),nom2 or f"Cliente {ced2[:6]}",plan2,int(dias2),fi2)
                    st.success(f"✅ {ced2} guardado — {plan2.upper()} {dias2} dias")
                else: st.error("Ingresa la cedula")
        with c2:
            st.markdown("### 👥 Lista de clientes")
            for u in db_todos():
                if u["cedula"]=="admin": continue
                ok2,p2,dr2=db_acceso(u["cedula"])
                st.markdown(f"**{u['nombre']}** `{u['cedula']}` {pll(u['plan'],dr2)}",unsafe_allow_html=True)

    # ── Tab 5: Escalera ────────────────────────────────────
    with tabs[4]:
        st.markdown("### 🏆 Reto Escalera del Dia")
        st.info("La escalera usa automaticamente los picks con mayor confianza Y edge positivo publicados hoy.")
        picks_hoy=db_picks_get(hoy_str,plan="admin")
        picks_reales=[p for p in picks_hoy if "🔒" not in str(p.get("mercado",""))]
        # Ordenar: primero por edge positivo, luego por confianza
        def esc_score(p):
            edge=p.get("edge") or -99
            conf=p.get("confianza") or 0
            return (1 if edge and edge>=2 else 0, conf)
        picks_esc=sorted(picks_reales, key=esc_score, reverse=True)[:8]
        if picks_esc:
            cuota_total=1.0
            for p in picks_esc:
                if p.get("cuota"): cuota_total=round(cuota_total*p["cuota"],2)
            st.markdown(f"**{len(picks_esc)} pasos en la escalera | Cuota combinada: {cuota_total}x**")
            for i,p in enumerate(picks_esc,1):
                edge_v=p.get("edge")
                edge_str=f"Edge: {edge_v:+.1f}%" if edge_v is not None else ""
                val_ico="🔥" if edge_v and edge_v>=5 else ("✅" if edge_v and edge_v>=2 else "⚪")
                cls="ap" if p.get("rango")=="A+" else "b"
                st.markdown(f'<div class="esc-box"><b>Paso {i}:</b> {p["local"]} vs {p["visitante"]} · <i>{p["liga"]}</i><br>{val_ico} <b>{p["mercado"]}</b> @ {p.get("cuota","?")} · Conf: {p.get("confianza","?")}% {edge_str}</div>',unsafe_allow_html=True)
        else:
            st.info("Publica picks primero para armar la escalera.")

    if st.button("🚪 Cerrar sesion",key="logout_admin"): st.session_state.clear(); st.rerun()

# ══════════════════════════════════════════════════════════
# PANTALLA GRATIS
# ══════════════════════════════════════════════════════════
def pantalla_gratis(u):
    hdr()
    ok,plan,dr=db_acceso(u["cedula"])
    st.markdown(f'👋 Hola **{u["nombre"]}** {pll("gratis",dr)} · Max 5 partidos/dia',unsafe_allow_html=True)
    t1,t2=st.tabs(["📁 Analizar archivo","📢 Picks del dia"])
    with t1:
        st.info("Plan gratuito: primeros 5 partidos, sin datos reales de API. Actualiza para acceso completo.")
        cons=db_consultas(u["cedula"])
        if cons>5:
            st.warning("Limite de 5 consultas diarias alcanzado. Vuelve manana o actualiza tu plan.")
        else:
            p=widget_archivo(max_p=5,key="free_up")
            if p and st.button("🦂 Analizar"):
                prg=st.progress(0,"Analizando...")
                res=analizar_lista(p,usar_api=False,prog=prg)
                prg.progress(1.0,"Listo ✅")
                for r in res:
                    st.markdown(f'{bdg(r["rango"])} **{r["local"]} vs {r["visitante"]}** · {r.get("hora","")} · **{r["mk2"]}** · xG:{r["xl"]}-{r["xv"]} · O2.5:{r["over25"]}% · Conf:{r["confianza"]}%',unsafe_allow_html=True)
                st.download_button("⬇️ Descargar Excel",data=exportar_excel(res,"Scorpion — Gratis"),
                    file_name=f"Scorpion_gratis_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with t2:
        picks=db_picks_get(str(date.today()),"gratis")
        if picks:
            for p in picks:
                cls="ap" if p.get("rango")=="A+" else "b"
                bl="🔒" in str(p.get("mercado",""))
                if bl:
                    st.markdown(f'<div class="pick-box"><b>{p["liga"]}</b> · {p["local"]} vs {p["visitante"]}<br><small>{p["mercado"]}</small></div>',unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="pick-box {cls}"><b>{p["liga"]}</b> · {p["local"]} vs {p["visitante"]} · {p.get("hora","")}<br>📌 <b>{p["mercado"]}</b> @ {p.get("cuota","?")} · Conf:{p.get("confianza","?")}%</div>',unsafe_allow_html=True)
        else:
            st.info("No hay picks publicados hoy.")
    if st.button("🚪 Cerrar sesion",key="logout_free"): st.session_state.clear(); st.rerun()

# ══════════════════════════════════════════════════════════
# PANTALLA PAGO
# ══════════════════════════════════════════════════════════
def pantalla_pago(u,plan):
    hdr()
    ok,pv,dr=db_acceso(u["cedula"])
    pl_lbl={"dia":"📅 Plan Dia","semana":"📆 Plan Semana","mes":"👑 Plan Mes"}.get(plan,plan)
    st.markdown(f'👋 Hola **{u["nombre"]}** {pll(plan,dr)}',unsafe_allow_html=True)
    st.markdown("---")
    tl=["🏟️ Por Liga","📁 Subir Archivo","📢 Picks","🏆 Escalera"]
    if plan=="mes": tl.append("🔗 Combinadas"); tl.append("📈 Historial")
    tabs=st.tabs(tl)

    with tabs[0]:
        st.markdown("### Selecciona liga y periodo")

        # Mostrar ligas activas hoy
        with st.expander("⚡ Ver ligas con partidos HOY", expanded=False):
            st.caption("Consultando API-Football para ligas activas hoy...")
            hoy_str2 = str(date.today())
            ligas_activas_hoy = []
            for ln, lid in list(LIGAS.items())[:8]:  # revisar primeras 8 para no agotar requests
                try:
                    fx_test = get_fx_dia(lid, hoy_str2)
                    if fx_test:
                        ligas_activas_hoy.append((ln, len(fx_test)))
                    time.sleep(0.3)
                except: pass
            if ligas_activas_hoy:
                for ln, cnt in ligas_activas_hoy:
                    st.markdown(f"✅ **{ln}** — {cnt} partidos hoy")
            else:
                st.info("No se encontraron partidos hoy en las ligas principales, o la API no responde.")

        c1,c2=st.columns([2,1])
        with c1:
            st.caption("Puedes seleccionar cualquier liga, este o no activa hoy.")
            if plan=="mes":
                ligas_sel=st.multiselect("Ligas",list(LIGAS.keys()),default=[list(LIGAS.keys())[0]])
            else:
                ligas_sel=[st.selectbox("Liga",list(LIGAS.keys()))]
        with c2:
            if plan=="dia":
                fecha_s=st.date_input("Fecha",value=date.today()); modo="dia"
            else:
                ops=["Hoy","Dia especifico","Esta semana","Semana personalizada"]
                if plan in ("semana","mes"): ops.append("Dias de la semana")
                ml=st.radio("Periodo",ops)
                if ml=="Hoy": fecha_s=date.today(); modo="dia"
                elif ml=="Dia especifico": fecha_s=st.date_input("Fecha"); modo="dia"
                elif ml=="Esta semana":
                    hoy=date.today(); lu=hoy-timedelta(days=hoy.weekday())
                    fd=lu; fh=lu+timedelta(days=6); modo="rango"
                elif ml=="Semana personalizada":
                    fd=st.date_input("Desde",value=date.today())
                    fh=st.date_input("Hasta",value=date.today()+timedelta(days=6)); modo="rango"
                elif ml=="Dias de la semana":
                    dias_n=st.multiselect("Dias",["Lunes","Martes","Miercoles","Jueves","Viernes","Sabado","Domingo"])
                    modo="dias"

        if st.button("🦂 Obtener y Analizar",key="btn_liga_pago"):
            todos=[]
            with st.spinner("Obteniendo fixtures de API-Football..."):
                for ln in ligas_sel:
                    lid=LIGAS[ln]
                    if modo=="dia": fx=get_fx_dia(lid,str(fecha_s))
                    elif modo=="rango": fx=get_fx_rango(lid,str(fd),str(fh))
                    elif modo=="dias":
                        fx=[]; hoy=date.today(); lu=hoy-timedelta(days=hoy.weekday())
                        mp={"Lunes":0,"Martes":1,"Miercoles":2,"Jueves":3,"Viernes":4,"Sabado":5,"Domingo":6}
                        for d in dias_n:
                            fd2=lu+timedelta(days=mp.get(d,0)); fx+=get_fx_dia(lid,str(fd2))
                    todos+=[fx2p(f) for f in fx]; time.sleep(0.3)
            if not todos:
                st.warning("No se encontraron partidos. La liga puede estar en receso o la temporada no ha comenzado.")
            else:
                prg=st.progress(0,"Analizando...")
                res=analizar_lista(todos,usar_api=True,prog=prg)
                prg.progress(1.0,"Listo ✅")
                ap=sum(1 for r in res if r["rango"]=="A+")
                b=sum(1 for r in res if r["rango"]=="B")
                api_ok=sum(1 for r in res if "API" in r.get("fuente_l",""))
                cp=round(sum(r.get("confianza",0) for r in res)/len(res)) if res else 0
                c1,c2,c3,c4=st.columns(4)
                for col,(v,l) in zip([c1,c2,c3,c4],[(len(res),"Partidos"),(ap,"Rango A+"),(api_ok,"Datos reales"),(f"{cp}%","Confianza prom")]):
                    with col: st.markdown(f'<div class="mc"><div class="v">{v}</div><div class="l">{l}</div></div>',unsafe_allow_html=True)
                st.markdown("### 🏆 Top Picks")
                top=sorted([r for r in res if r["rango"] in ("A+","B")],
                           key=lambda x:(x.get("confianza",0),max(x.get("p1",0),x.get("p2",0))),reverse=True)[:8]
                for r in top:
                    st.markdown(f'{bdg(r["rango"])} **{r["local"]} vs {r["visitante"]}** ({r["liga"][:16]}) · {r["hora"]} · **{r["mk2"]}** · xG:{r["xl"]}-{r["xv"]} · O2.5:{r["over25"]}% · Tiros:~{r.get("tiros_tot",0)} · Conf:{r["confianza"]}%',unsafe_allow_html=True)
                if top:
                    st.markdown("### 🔬 Detalle modelos")
                    for r in top[:3]:
                        with st.expander(f"{r['local']} vs {r['visitante']} — {r['rango']} ({r['confianza']}% conf)"):
                            dc1,dc2,dc3,dc4=st.columns(4)
                            with dc1: st.metric("Poisson",f'{r.get("p1_po",0):.1f}%')
                            with dc2: st.metric("Dixon-Coles",f'{r.get("p1_dc",0):.1f}%')
                            with dc3: st.metric("Monte Carlo",f'{r.get("p1_mc",0):.1f}%')
                            with dc4: st.metric("Elo",f'{r.get("p1_el",0):.1f}%')
                            st.markdown(f"xG: {r['xl']} vs {r['xv']} | Total: {r['xt']}")
                            st.markdown(f"Tiros: Local ~{r.get('tiros_l',0)} | Visita ~{r.get('tiros_v',0)} | Total ~{r.get('tiros_tot',0)}")
                            st.markdown(f"Corners: {r.get('corners_str','')} | Tarjetas: {r.get('tar_str','')}")
                            st.markdown(f"Marcadores: {' | '.join([f'{k}({v}%)' for k,v in list(r.get('top_ex',{}).items())[:5]])}")
                fl=str(fecha_s) if modo=="dia" else f"{fd}_al_{fh}"
                xl=exportar_excel(res,f"Scorpion Elite — {pl_lbl}")
                st.download_button("⬇️ Descargar Excel completo",data=xl,
                    file_name=f"ScorpionElite_{fl}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                for r in res: db_historial_guardar(u["cedula"],r,r)

    with tabs[1]:
        st.info("Sube una captura de Sofascore, FlashScore o cualquier app. La IA lee los partidos automaticamente.")
        p=widget_archivo(key="pago_up")
        if p and st.button("🦂 Analizar archivo",key="btn_arch"):
            prg=st.progress(0,"Analizando...")
            res=analizar_lista(p,usar_api=True,prog=prg)
            prg.progress(1.0,"Listo ✅")
            xl=exportar_excel(res,"Scorpion Elite — Archivo")
            st.download_button("⬇️ Descargar Excel",data=xl,
                file_name=f"Scorpion_custom_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tabs[2]:
        picks=db_picks_get(str(date.today()),plan)
        if picks:
            for p in picks:
                cls="ap" if p.get("rango")=="A+" else "b"
                bl="🔒" in str(p.get("mercado",""))
                if bl:
                    st.markdown(f'<div class="pick-box"><b>{p["liga"]}</b> · {p["local"]} vs {p["visitante"]}<br><small>{p["mercado"]}</small></div>',unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="pick-box {cls}"><b>{p["liga"]}</b> · {p["local"]} vs {p["visitante"]} · {p.get("hora","")}<br>📌 <b>{p["mercado"]}</b> @ {p.get("cuota","?")} · Edge:{p.get("edge","?")}% · Conf:{p.get("confianza","?")}%{"<br><small>"+p["notas"]+"</small>" if p.get("notas") else ""}</div>',unsafe_allow_html=True)
        else:
            st.info("No hay picks publicados para hoy. El administrador los publica diariamente.")

    with tabs[3]:
        st.markdown("### 🏆 Reto Escalera")
        st.caption("Picks del dia ordenados por mayor confianza y valor estadistico.")
        picks_e=db_picks_get(str(date.today()),plan)
        reales=[p for p in picks_e if "🔒" not in str(p.get("mercado",""))]
        def esc_score(p):
            edge=p.get("edge") or -99
            conf=p.get("confianza") or 0
            return (1 if edge and edge>=2 else 0, conf)
        picks_esc=sorted(reales, key=esc_score, reverse=True)[:8]
        if picks_esc:
            cuota_total=1.0
            for p in picks_esc:
                if p.get("cuota"): cuota_total=round(cuota_total*p["cuota"],2)
            st.markdown(f"**{len(picks_esc)} pasos | Cuota combinada: {cuota_total}x**")
            for i,p in enumerate(picks_esc,1):
                edge_v=p.get("edge")
                edge_str=f" · Edge: {edge_v:+.1f}%" if edge_v is not None else ""
                val_ico="🔥" if edge_v and edge_v>=5 else ("✅" if edge_v and edge_v>=2 else "⚪")
                cls="ap" if p.get("rango")=="A+" else "b"
                st.markdown(f'<div class="esc-box"><b>Paso {i}:</b> {p["local"]} vs {p["visitante"]} · {p["liga"]}<br>{val_ico} <b>{p["mercado"]}</b> @ {p.get("cuota","?")} · Conf: {p.get("confianza","?")}%{edge_str}</div>',unsafe_allow_html=True)
        else:
            st.info("No hay picks publicados hoy para la escalera.")

    if plan=="mes" and len(tabs)>4:
        with tabs[4]:
            st.markdown("### 🔗 Combinadas del Dia")
            st.markdown("""
            **¿Cómo funciona una combinada?**
            Una combinada es apostar a varios partidos a la vez. Para ganar, **todos** los picks deben acertar.
            La cuota total se multiplica (ej: 1.8 × 2.0 × 1.5 = 5.4x tu apuesta).
            Cuanto mas picks agregas, mayor es la cuota pero menor la probabilidad de acertar.
            """)

            picks_c=db_picks_get(str(date.today()),plan)
            reales_c=[p for p in picks_c if "🔒" not in str(p.get("mercado","")) and p.get("cuota")]
            if reales_c:
                st.markdown("**Selecciona los picks que quieres combinar:**")
                sel_comb=[]
                for p in reales_c:
                    conf=p.get("confianza",0); cuota_p=p.get("cuota",1)
                    edge=p.get("edge",0)
                    edge_str=f" · Edge: {edge:+.1f}%" if edge else ""
                    lbl=f'{p["local"]} vs {p["visitante"]} — {p["mercado"]} @ {cuota_p} · Conf: {conf}%{edge_str}'
                    if st.checkbox(lbl, key=f"comb_{p['id']}"):
                        sel_comb.append(p)

                if sel_comb:
                    cuota_c=1.0; prob_c=1.0
                    for s in sel_comb:
                        if s.get("cuota"):     cuota_c=round(cuota_c*float(s["cuota"]),2)
                        if s.get("confianza"): prob_c=round(prob_c*float(s["confianza"])/100,5)
                    prob_pct=round(prob_c*100,1)
                    # Kelly conservador (1/4)
                    b=cuota_c-1
                    kelly_f=max(0,(prob_c*b-(1-prob_c))/b)*0.25 if b>0 else 0
                    kelly_pct=round(kelly_f*100,1)
                    # Ganancia neta si acierta con $100 base
                    ganancia_100=round(100*cuota_c-100,1)

                    st.markdown("---")
                    st.markdown("### 📊 Tu Combinada")

                    # Tabla de picks seleccionados
                    tbl_comb='<table style="width:100%;border-collapse:collapse;font-size:.83rem;margin:8px 0"><thead><tr style="background:#0a0a1e;color:#ffd700"><th style="padding:6px 12px;text-align:left">Partido</th><th style="padding:6px 12px;text-align:left">Pick</th><th style="padding:6px 12px;text-align:center">Cuota</th><th style="padding:6px 12px;text-align:center">Confianza</th></tr></thead><tbody>'
                    for i_c,s in enumerate(sel_comb):
                        bg_c="#0d0d18" if i_c%2==0 else "#0a0a1e"
                        tbl_comb+=f'<tr style="background:{bg_c}"><td style="padding:5px 12px;color:#ccc">{s["local"][:12]} vs {s["visitante"][:12]}</td><td style="padding:5px 12px;color:#ffd700;font-weight:600">{s["mercado"]}</td><td style="padding:5px 12px;text-align:center;color:#fff">{s.get("cuota","?")}</td><td style="padding:5px 12px;text-align:center;color:#00ee66">{s.get("confianza","?")}%</td></tr>'
                    tbl_comb+="</tbody></table>"
                    st.markdown(tbl_comb, unsafe_allow_html=True)

                    # Resumen de metricas con explicacion
                    tbl_res=f'''<table style="width:100%;border-collapse:collapse;font-size:.85rem;margin:10px 0">
                      <thead><tr style="background:#0a1e0a;color:#ffd700">
                        <th style="padding:7px 14px;text-align:left">Metrica</th>
                        <th style="padding:7px 14px;text-align:center">Valor</th>
                        <th style="padding:7px 14px;text-align:left">Que significa</th>
                      </tr></thead><tbody>
                      <tr style="background:#0d1e0d">
                        <td style="padding:6px 14px;color:#aaa">Picks en la combinada</td>
                        <td style="padding:6px 14px;text-align:center;color:#fff;font-weight:700">{len(sel_comb)}</td>
                        <td style="padding:6px 14px;color:#888">Todos deben acertar para ganar</td>
                      </tr>
                      <tr style="background:#0a1a0a">
                        <td style="padding:6px 14px;color:#aaa">Cuota total combinada</td>
                        <td style="padding:6px 14px;text-align:center;color:#ffd700;font-weight:700">{cuota_c}x</td>
                        <td style="padding:6px 14px;color:#888">Multiplica tu apuesta por {cuota_c}</td>
                      </tr>
                      <tr style="background:#0d1e0d">
                        <td style="padding:6px 14px;color:#aaa">Probabilidad estimada</td>
                        <td style="padding:6px 14px;text-align:center;color:{"#00ee66" if prob_pct>=40 else "#ffaa00" if prob_pct>=20 else "#ff4444"};font-weight:700">{prob_pct}%</td>
                        <td style="padding:6px 14px;color:#888">{"Razonable ✅" if prob_pct>=40 else "Arriesgada ⚠️" if prob_pct>=20 else "Muy arriesgada ❌"}</td>
                      </tr>
                      <tr style="background:#0a1a0a">
                        <td style="padding:6px 14px;color:#aaa">Ganancia neta por $100</td>
                        <td style="padding:6px 14px;text-align:center;color:#00ee66;font-weight:700">${ganancia_100}</td>
                        <td style="padding:6px 14px;color:#888">Si apuestas $100 y aciertas, ganas ${ganancia_100} de utilidad</td>
                      </tr>
                      <tr style="background:#0d1e0d">
                        <td style="padding:6px 14px;color:#aaa">Kelly recomendado</td>
                        <td style="padding:6px 14px;text-align:center;color:#ffd700;font-weight:700">{kelly_pct}% del bankroll</td>
                        <td style="padding:6px 14px;color:#888">Si tienes $1000, apuesta maximo ${round(kelly_pct*10,0)}</td>
                      </tr>
                    </tbody></table>'''
                    st.markdown(tbl_res, unsafe_allow_html=True)

                    if prob_pct < 20:
                        st.error("⚠️ Esta combinada tiene menos de 20% de probabilidad. Es muy arriesgada. Considera reducir el numero de picks.")
                    elif prob_pct < 40:
                        st.warning("⚠️ Combinada arriesgada. Apuesta solo una fraccion pequena del bankroll.")
                    else:
                        st.success(f"✅ Combinada razonable con {prob_pct}% de probabilidad estimada.")
            else:
                st.info("No hay picks publicados hoy para combinar. El administrador debe publicar picks primero.")

        with tabs[5]:
            st.markdown("### 📈 Mi Historial")
            hist=db_historial_get(u["cedula"],30)
            if hist:
                for h in hist:
                    st.markdown(f'**{h["local"]} vs {h["visitante"]}** · {h["liga"]} · {h["fecha"]} · {h["mercado"]} · {h["rango"]}')
            else:
                st.info("No tienes analisis guardados aun.")

    if st.button("🚪 Cerrar sesion",key="logout_pago"): st.session_state.clear(); st.rerun()
    st.markdown('<div class="ft">🦂 Scorpion Elite V4 Pro 2025 · Solo uso informativo</div>',unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# ROUTER
# ══════════════════════════════════════════════════════════
init_db()
if "li" not in st.session_state: st.session_state.li=False
if not st.session_state.li:
    pantalla_login()
else:
    u=st.session_state.get("u",{}); ced=st.session_state.get("ced","")
    ok,plan,dr=db_acceso(ced)
    if plan=="admin": pantalla_admin()
    elif plan=="gratis": pantalla_gratis(u)
    elif plan in ("dia","semana","mes"): pantalla_pago(u,plan)
    else:
        st.error("Acceso vencido. Contacta al administrador.")
        if st.button("Volver"): st.session_state.clear(); st.rerun()
